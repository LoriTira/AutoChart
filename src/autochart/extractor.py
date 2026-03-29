"""Auto-extraction of configuration metadata from input Excel workbooks.

Scans INPUT sheets in an Excel workbook and extracts disease name, year range,
rate unit, data source, geography, demographics, and reference group so users
don't have to type configuration manually.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from autochart.config import ChartConfig


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class ExtractedConfig:
    disease_name: str | None = None
    years: str | None = None
    rate_unit: str | None = None
    rate_denominator: int | None = None
    data_source: str | None = None
    geography: str | None = None
    demographics: list[str] | None = None
    reference_group: str | None = None
    confidence: dict[str, float] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_DISEASE_RE = re.compile(
    r"([\w\s]*(?:Mortality|Hospitalizations?|Morbidity|Incidence))"
)
_YEAR_RE = re.compile(r"(\d{4})\s*[-\u2013]\s*(\d{4})")
_RATE_RE = re.compile(r"[Pp]er\s+([\d,]+)")
_DATA_SOURCE_RE = re.compile(r"DATA\s*SOURCE\s*:?\s*(.+)", re.IGNORECASE | re.DOTALL)

_KNOWN_RACE_LABELS = {
    "Asian", "Black", "Latinx", "Latino", "Hispanic",
    "White", "Native", "Pacific",
}

# Common abbreviations found in workbook titles
_DISEASE_ALIASES: dict[str, str] = {
    "cerebro": "Cerebrovascular",
}


def _cell_texts(ws, max_row: int = 15):
    """Yield (row_number, str) for every non-empty text cell in the first *max_row* rows."""
    for i, row in enumerate(ws.iter_rows(max_row=max_row, values_only=True), 1):
        for val in row:
            if isinstance(val, str) and val.strip():
                yield i, val.strip()


def _normalize_disease(raw: str) -> str:
    """Clean up a raw disease-keyword match into a canonical disease name."""
    # Remove surrounding quotes
    cleaned = raw.replace("'", "").replace("\u2018", "").replace("\u2019", "")
    # Strip leading filler words like "All" or "Boston"
    cleaned = re.sub(r"^(?:All|Boston)\s+", "", cleaned).strip()
    # Expand known abbreviations
    for abbr, full in _DISEASE_ALIASES.items():
        cleaned = re.sub(rf"(?i)\b{abbr}\b", full, cleaned)
    return cleaned.strip()


def _normalize_race(label: str) -> str | None:
    """Map a raw race cell value to a canonical label, or *None* if not a race."""
    cleaned = label.replace("_", " ").replace("'", "").strip()
    # Remove common suffixes like "nL"
    cleaned = re.sub(r"\s*nL$", "", cleaned, flags=re.IGNORECASE).strip()
    for known in _KNOWN_RACE_LABELS:
        if cleaned.lower() == known.lower():
            return known
    return None


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------

def extract_config(
    path: str | Path,
    sheets: list[str] | None = None,
) -> ExtractedConfig:
    """Open *path* read-only and extract as much config as possible.

    Iterates over every sheet whose name starts with ``INPUT`` (or the
    explicit *sheets* list, if given), scanning the first 15 rows of each
    for metadata.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    disease_counts: Counter[str] = Counter()
    year_counts: Counter[str] = Counter()
    rate_counts: Counter[int] = Counter()
    data_sources: list[str] = []
    geography_found = False
    race_labels: set[str] = set()

    if sheets is not None:
        input_sheets = sheets
    else:
        input_sheets = [name for name in wb.sheetnames if name.upper().startswith("INPUT")]

    for sheet_name in input_sheets:
        ws = wb[sheet_name]

        for row_num, text in _cell_texts(ws, max_row=15):
            # Normalise non-breaking spaces for all matching
            text_clean = text.replace("\xa0", " ")

            # --- Disease name ---
            # Strip single-quotes before matching so "'Cancer' Mortality" works
            text_unquoted = text_clean.replace("'", "")
            for m in _DISEASE_RE.finditer(text_unquoted):
                disease = _normalize_disease(m.group(1))
                if disease:
                    disease_counts[disease] += 1

            # --- Years ---
            ym = _YEAR_RE.search(text_clean)
            if ym:
                year_counts[f"{ym.group(1)}-{ym.group(2)}"] += 1

            # --- Rate unit ---
            rm = _RATE_RE.search(text_clean)
            if rm:
                num_str = rm.group(1).replace(",", "")
                try:
                    rate_counts[int(num_str)] += 1
                except ValueError:
                    pass

            # --- Data source ---
            dm = _DATA_SOURCE_RE.search(text_clean)
            if dm:
                source_text = dm.group(0).strip()
                # Collapse internal whitespace / newlines
                source_text = re.sub(r"\s+", " ", source_text)
                data_sources.append(source_text)

            # --- Geography ---
            if row_num <= 5 and "Boston" in text_clean:
                geography_found = True

            # --- Demographics (race labels) ---
            normalized = _normalize_race(text_clean)
            if normalized:
                race_labels.add(normalized)

    wb.close()

    # ---- Aggregate results ----
    result = ExtractedConfig()

    # Disease name: most common match
    if disease_counts:
        result.disease_name = disease_counts.most_common(1)[0][0]
        n_sheets = disease_counts.most_common(1)[0][1]
        result.confidence["disease_name"] = 0.9 if n_sheets >= 2 else 0.6

    # Years
    if year_counts:
        result.years = year_counts.most_common(1)[0][0]
        result.confidence["years"] = 0.95

    # Rate unit & denominator
    if rate_counts:
        denom = rate_counts.most_common(1)[0][0]
        result.rate_denominator = denom
        result.rate_unit = f"per {denom:,} residents"
        result.confidence["rate_unit"] = 0.9

    # Data source
    if data_sources:
        result.data_source = data_sources[0]
        result.confidence["data_source"] = 0.95

    # Geography
    if geography_found:
        result.geography = "Boston"
        result.confidence["geography"] = 0.8

    # Demographics
    demographics = sorted(race_labels - {None})
    if demographics:
        result.demographics = demographics
        result.confidence["demographics"] = 0.9 if len(demographics) >= 3 else 0.6

    # Reference group
    if demographics and "White" in demographics:
        result.reference_group = "White"
        result.confidence["reference_group"] = 0.7

    return result


# ---------------------------------------------------------------------------
# Merge function
# ---------------------------------------------------------------------------

def build_config(
    extracted: ExtractedConfig,
    overrides: dict | None = None,
) -> ChartConfig:
    """Build a :class:`ChartConfig` from extracted metadata plus optional user overrides.

    Priority: user overrides > extracted values > hardcoded defaults.
    """
    if overrides is None:
        overrides = {}

    disease_name = overrides.get("disease_name") or extracted.disease_name
    years = overrides.get("years") or extracted.years

    if not disease_name:
        raise ValueError(
            "Could not auto-detect disease name. Please provide --disease."
        )
    if not years:
        raise ValueError(
            "Could not auto-detect year range. Please provide --years."
        )

    rate_denominator = (
        overrides.get("rate_denominator")
        or extracted.rate_denominator
        or 100000
    )
    rate_unit = (
        overrides.get("rate_unit")
        or extracted.rate_unit
        or f"per {rate_denominator:,} residents"
    )
    data_source = overrides.get("data_source") or extracted.data_source or ""
    geography = overrides.get("geography") or extracted.geography or "Boston"
    demographics = (
        overrides.get("demographics")
        or extracted.demographics
        or ["Asian", "Black", "Latinx", "White"]
    )
    reference_group = (
        overrides.get("reference_group")
        or extracted.reference_group
        or "White"
    )

    return ChartConfig(
        disease_name=disease_name,
        rate_unit=rate_unit,
        rate_denominator=rate_denominator,
        data_source=data_source,
        years=years,
        demographics=demographics,
        reference_group=reference_group,
        geography=geography,
    )
