"""Manifest-driven template builder.

Uses template packages (template.xlsx + manifest.json) to produce
publication-ready output:

  1. Opens template .xlsx and fills data cells (charts auto-update)
  2. Applies OOXML post-processing (Montserrat, pattern fills, asterisks)
  3. Injects floating text boxes (descriptive paragraphs + footnotes)
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from autochart.config import (
    ChartConfig,
    ChartSetAData,
    ChartSetBData,
    ChartSetCData,
    ChartSetType,
    Part3Data,
    RateComparison,
)
from autochart.builder.postprocess import ChartPatch, postprocess_xlsx
from autochart.builder.textbox_updater import (
    TextBoxSpec,
    TextParagraph,
    TextRun,
    inject_text_boxes,
    make_description_textbox,
    make_footnote_textbox,
)
from autochart.template_packages.loader import (
    TemplateBlock,
    TemplatePackage,
    TextBoxAnchor,
    get_template,
    get_template_by_type,
)
from autochart.text.generator import TextGenerator


# ---------------------------------------------------------------------------
# Data filling
# ---------------------------------------------------------------------------

def _fill_set_a(ws, block: TemplateBlock, data: ChartSetAData, config: ChartConfig) -> None:
    """Fill a Set A block (race vs rest of city)."""
    if block.race_cells:
        for rc in block.race_cells:
            ws[rc] = data.race_name
    if block.label_cell:
        ws[block.label_cell] = f"All {config.disease_name}"

    values = [
        data.boston.group_rate, data.boston.reference_rate, data.boston_overall_rate,
        data.female.group_rate, data.female.reference_rate, data.female_overall_rate,
        data.male.group_rate, data.male.reference_rate, data.male_overall_rate,
    ]
    for dc, val in zip(block.data_cells, values):
        ws[dc] = val

    ws[block.title_cell] = (
        f"{config.disease_name}\u2020 for {data.race_name} Residents, {config.years}"
    )


def _fill_set_b(ws, block: TemplateBlock, data: ChartSetBData, config: ChartConfig) -> None:
    """Fill a Set B block (race vs reference group)."""
    if block.race_cell:
        ws[block.race_cell] = data.race_name

    values = [data.comparison.group_rate, data.comparison.reference_rate, data.boston_overall_rate]
    for dc, val in zip(block.data_cells, values):
        ws[dc] = val

    ws[block.title_cell] = (
        f"{config.disease_name}\u2020, {data.race_name} Residents "
        f"Compared to {config.reference_group} Residents, {config.years}"
    )


def _fill_set_c(ws, block: TemplateBlock, data: ChartSetCData, config: ChartConfig) -> None:
    """Fill a Set C block (all races combined)."""
    if block.header_cells:
        labels = [c.group_name for c in data.comparisons] + [
            config.reference_group, config.geography
        ]
        for hc, label in zip(block.header_cells, labels):
            ws[hc] = label

    rates = [c.group_rate for c in data.comparisons]
    values = rates + [data.comparisons[0].reference_rate, data.boston_overall_rate]
    for dc, val in zip(block.data_cells, values):
        ws[dc] = val

    ws[block.title_cell] = f"{config.disease_name}\u2020 by Race, {config.years}"


def _fill_part3(ws, block: TemplateBlock, data: Part3Data, config: ChartConfig) -> None:
    """Fill a Part 3 block (gender x race)."""
    if block.race_cells:
        race_names = [c.group_name for c in data.female_comparisons]
        for rc, name in zip(block.race_cells, race_names + race_names):
            ws[rc] = name

    f_rates = [c.group_rate for c in data.female_comparisons]
    m_rates = [c.group_rate for c in data.male_comparisons]
    values = (
        f_rates + [data.female_comparisons[0].reference_rate, data.female_boston_rate]
        + m_rates + [data.male_comparisons[0].reference_rate, data.male_boston_rate]
    )
    for dc, val in zip(block.data_cells, values):
        ws[dc] = val

    ws[block.title_cell] = (
        f"{config.disease_name}\u2020 by Sex and Race, {config.years}"
    )


_FILLERS = {
    ChartSetType.A: _fill_set_a,
    ChartSetType.B: _fill_set_b,
    ChartSetType.C: _fill_set_c,
    ChartSetType.PART_3: _fill_part3,
}


# ---------------------------------------------------------------------------
# Chart patch computation
# ---------------------------------------------------------------------------

def _compute_patches(
    pkg: TemplatePackage,
    data_list: list,
    config: ChartConfig,
) -> list[ChartPatch]:
    """Compute ChartPatch objects for post-processing."""
    patches = []
    for i, block in enumerate(pkg.blocks):
        if i >= len(data_list):
            break

        asterisk_points = []
        data = data_list[i]

        if pkg.chart_set_type == ChartSetType.A:
            # Set A: check significance for each of the 3 groups (Boston, Female, Male)
            # Bars at indices 0,3,6 are the race bars
            for idx, comp in [(0, data.boston), (3, data.female), (6, data.male)]:
                if comp.is_significant:
                    asterisk_points.append(idx)

        elif pkg.chart_set_type == ChartSetType.B:
            # Set B: check if race vs reference is significant (bar 0)
            if data.comparison.is_significant:
                asterisk_points.append(0)

        elif pkg.chart_set_type == ChartSetType.C:
            # Set C: check each race's comparison
            for j, comp in enumerate(data.comparisons):
                if comp.is_significant:
                    asterisk_points.append(j)

        elif pkg.chart_set_type == ChartSetType.PART_3:
            # Part 3: check female (0-2) and male (5-7) race comparisons
            for j, comp in enumerate(data.female_comparisons):
                if comp.is_significant:
                    asterisk_points.append(j)
            for j, comp in enumerate(data.male_comparisons):
                if comp.is_significant:
                    asterisk_points.append(j + 5)

        patches.append(ChartPatch(
            chart_index=block.chart_index,
            pattern_fill_points=list(block.pattern_fill_points),
            asterisk_points=asterisk_points,
        ))

    return patches


# ---------------------------------------------------------------------------
# Text box generation
# ---------------------------------------------------------------------------

def _build_text_boxes(
    pkg: TemplatePackage,
    data_list: list,
    config: ChartConfig,
) -> list[TextBoxSpec]:
    """Build text box specs for all blocks in a template."""
    text_gen = TextGenerator(config)
    text_boxes = []

    for i, block in enumerate(pkg.blocks):
        if i >= len(data_list):
            break

        data = data_list[i]

        # Descriptive text
        if "description" in block.text_boxes:
            anchor = block.text_boxes["description"]

            if pkg.chart_set_type == ChartSetType.A:
                desc = text_gen.descriptive_text_set_a(data)
            elif pkg.chart_set_type == ChartSetType.B:
                desc = text_gen.descriptive_text_set_b(data)
            elif pkg.chart_set_type == ChartSetType.C:
                desc = text_gen.descriptive_text_set_c(data)
            elif pkg.chart_set_type == ChartSetType.PART_3:
                desc = text_gen.descriptive_text_part3(data)
            else:
                desc = ""

            if desc:
                text_boxes.append(make_description_textbox(
                    anchor={"from_col": anchor.from_col, "from_row": anchor.from_row,
                            "to_col": anchor.to_col, "to_row": anchor.to_row},
                    text=desc,
                    font_name="Calibri",
                    font_size=10.0,
                ))

        # Footnote
        if "footnote" in block.text_boxes:
            anchor = block.text_boxes["footnote"]
            footnote_text = text_gen.footnote()
            footnote_lines = footnote_text.split("\n")

            text_boxes.append(make_footnote_textbox(
                anchor={"from_col": anchor.from_col, "from_row": anchor.from_row,
                        "to_col": anchor.to_col, "to_row": anchor.to_row},
                lines=footnote_lines,
                font_name="Calibri",
                font_size=9.0,
            ))

    return text_boxes


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

@dataclass
class TableAssignment:
    """Assigns a detected table to a template for generation."""
    template_id: str
    data_list: list[Any]
    config: ChartConfig


class TemplateBuilder:
    """Manifest-driven template builder with full pipeline."""

    def build_from_assignments(
        self,
        assignments: list[TableAssignment],
    ) -> dict[str, bytes]:
        """Build output files from a list of table-to-template assignments.

        Produces one .xlsx per assignment (each template gets its own file).

        Returns
        -------
        dict[str, bytes]
            Mapping from label to .xlsx bytes.
            Labels are "{disease} - {template_name}".
        """
        results = {}
        for a in assignments:
            pkg = get_template(a.template_id)
            label = f"{a.config.disease_name} - {pkg.name}"
            results[label] = self._build_one(pkg, a.data_list, a.config)
        return results

    def build_combined(
        self,
        assignments: list[TableAssignment],
    ) -> bytes:
        """Build a single .xlsx with one sheet per assignment.

        Each assignment becomes a sheet named after its template.
        Charts, text boxes, and formatting are all preserved.
        """
        from autochart.builder.combiner import combine_workbooks

        sheet_list: list[tuple[str, bytes]] = []
        used_names: set[str] = set()
        for a in assignments:
            pkg = get_template(a.template_id)
            xlsx_bytes = self._build_one(pkg, a.data_list, a.config)

            # Build a unique sheet name (max 31 chars for Excel)
            disease_short = a.config.disease_name[:15].strip()
            type_label = pkg.chart_set_type.value  # A, B, C, PART_3
            name = f"{disease_short} - {type_label}"
            # Ensure uniqueness
            if name in used_names:
                i = 2
                while f"{name} ({i})" in used_names:
                    i += 1
                name = f"{name} ({i})"
            # Excel sheet names max 31 chars
            name = name[:31]
            used_names.add(name)

            sheet_list.append((name, xlsx_bytes))

        return combine_workbooks(sheet_list)

    def build_disease(
        self,
        disease_name: str,
        tables: dict[ChartSetType, tuple[ChartConfig, list]],
    ) -> dict[str, bytes]:
        """Build output workbooks for a single disease.

        Produces one .xlsx per chart type. Returns {label: xlsx_bytes}.
        """
        results = {}
        for ct, (config, data_list) in tables.items():
            try:
                pkg = get_template_by_type(ct)
            except KeyError:
                continue
            label = f"{disease_name} - {pkg.name}"
            results[label] = self._build_one(pkg, data_list, config)
        return results

    def build_multi(
        self,
        disease_tables: dict[str, dict[ChartSetType, tuple[ChartConfig, list]]],
    ) -> dict[str, bytes]:
        """Build output files for multiple diseases. Returns {label: xlsx_bytes}."""
        results = {}
        for disease_name, tables in disease_tables.items():
            results.update(self.build_disease(disease_name, tables))
        return results

    def build_auto(
        self,
        sheet_results: list,
        requested_types: list[ChartSetType] | None = None,
    ) -> dict[str, bytes]:
        """Auto-build from parsed sheet results. Returns {label: xlsx_bytes}."""
        if requested_types is None:
            requested_types = list(ChartSetType)

        disease_tables: dict[str, dict[ChartSetType, tuple[Any, list]]] = {}
        for sr in sheet_results:
            d = sr.config.disease_name
            if d not in disease_tables:
                disease_tables[d] = {}
            for ct, data_list in sr.by_type.items():
                if ct not in requested_types:
                    continue
                if ct not in disease_tables[d]:
                    disease_tables[d][ct] = (sr.config, [])
                disease_tables[d][ct][1].extend(data_list)

        return self.build_multi(disease_tables)

    # -----------------------------------------------------------------------
    # Internal
    # -----------------------------------------------------------------------

    def _build_one(
        self,
        pkg: TemplatePackage,
        data_list: list,
        config: ChartConfig,
    ) -> bytes:
        """Build a single output file from one template + data.

        Steps:
          1. Open template and fill data cells (charts auto-update)
          2. Save via openpyxl
          3. Apply OOXML post-processing (Montserrat, patterns, asterisks)
          4. Inject text boxes (descriptive text + footnotes)
        """
        # Step 1: Fill data cells
        template_bytes = pkg.template_path.read_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb[pkg.sheet_name]

        filler = _FILLERS.get(pkg.chart_set_type)
        if filler:
            for i, block in enumerate(pkg.blocks):
                if i < len(data_list):
                    filler(ws, block, data_list[i], config)

        # Step 2: Save
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        # Step 3: Post-process (Montserrat font, pattern fills, asterisks)
        patches = _compute_patches(pkg, data_list, config)
        xlsx_bytes = postprocess_xlsx(xlsx_bytes, patches)

        # Step 4: Inject text boxes
        text_boxes = _build_text_boxes(pkg, data_list, config)
        if text_boxes:
            xlsx_bytes = inject_text_boxes(xlsx_bytes, pkg.sheet_name, text_boxes)

        return xlsx_bytes
