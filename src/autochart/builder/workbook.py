"""Assemble the complete AutoChart output workbook.

Uses **openpyxl** for cell data / formatting and the
:mod:`autochart.builder.injector` module for embedding charts and
text-box shapes that openpyxl cannot produce on its own.

The chart-creation methods (``add_chart_set_a``, etc.) are stubbed here
and will be fleshed out in later build steps.
"""

from __future__ import annotations

import io
from typing import Any, Sequence

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from autochart.config import (
    ChartConfig,
    ChartSetAData,
    ChartSetBData,
    ChartSetCData,
    ChartSetType,
    Part3Data,
)

# ---------------------------------------------------------------------------
# Shared style constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Aptos Narrow", size=11, bold=True)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_DATA_FONT = Font(name="Calibri", size=12)
_DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_HIGHLIGHT_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")


# ---------------------------------------------------------------------------
# WorkbookBuilder
# ---------------------------------------------------------------------------

class WorkbookBuilder:
    """High-level builder that assembles the final ``.xlsx`` output.

    Parameters
    ----------
    config:
        The :class:`ChartConfig` governing disease name, rate units,
        colours, demographics, etc.
    """

    def __init__(self, config: ChartConfig) -> None:
        self.config = config
        self.wb = openpyxl.Workbook()
        # Remove the default blank sheet created by openpyxl.
        self.wb.remove(self.wb.active)

    # ------------------------------------------------------------------
    # Chart-set stubs (to be implemented in Steps 4-6)
    # ------------------------------------------------------------------

    def _unique_sheet_name(self, base: str) -> str:
        """Return *base* if it doesn't exist yet, otherwise append ``_2``, ``_3``, etc."""
        if base not in self.wb.sheetnames:
            return base
        i = 2
        while f"{base}_{i}" in self.wb.sheetnames:
            i += 1
        return f"{base}_{i}"

    def add_chart_set_a(
        self, data_list: list[ChartSetAData], config: ChartConfig | None = None,
    ) -> None:
        """Add Chart Set A sheet with race charts.

        Each item in *data_list* represents one race group compared
        against the rest of Boston.

        Parameters
        ----------
        config:
            Optional per-sheet config override.  Falls back to ``self.config``.
        """
        if not data_list:
            return
        cfg = config or self.config
        ws = self.wb.create_sheet(self._unique_sheet_name("OUTPUT-1"))
        from autochart.charts.chart_set_a import build_chart_set_a_sheet
        build_chart_set_a_sheet(ws, data_list, cfg)

    def add_chart_set_b(
        self, data_list: list[ChartSetBData], config: ChartConfig | None = None,
    ) -> None:
        """Add Chart Set B sheet with race charts.

        Each item in *data_list* represents one race group compared
        against White.

        Parameters
        ----------
        config:
            Optional per-sheet config override.  Falls back to ``self.config``.
        """
        if not data_list:
            return
        cfg = config or self.config
        ws = self.wb.create_sheet(self._unique_sheet_name("OUTPUT-2"))
        from autochart.charts.chart_set_b import build_chart_set_b_sheet
        build_chart_set_b_sheet(ws, data_list, cfg)

    def add_chart_set_c(
        self, data: ChartSetCData, config: ChartConfig | None = None,
    ) -> None:
        """Add Chart Set C sheet with combined comparison chart.

        Parameters
        ----------
        config:
            Optional per-sheet config override.  Falls back to ``self.config``.
        """
        cfg = config or self.config
        ws = self.wb.create_sheet(self._unique_sheet_name("OUTPUT-3"))
        from autochart.charts.chart_set_c import build_chart_set_c_sheet
        build_chart_set_c_sheet(ws, data, cfg)

    def add_part_3(
        self, data: Part3Data, config: ChartConfig | None = None,
    ) -> None:
        """Add Part 3 sheet with gender x race chart.

        Parameters
        ----------
        config:
            Optional per-sheet config override.  Falls back to ``self.config``.
        """
        cfg = config or self.config
        ws = self.wb.create_sheet(self._unique_sheet_name("OUTPUT-4"))
        from autochart.charts.part_3 import build_part_3_sheet
        build_part_3_sheet(ws, data, cfg)

    # ------------------------------------------------------------------
    # Persistence
    # ------------------------------------------------------------------

    def save(self, path: str) -> None:
        """Save the workbook to a file on disk."""
        self.wb.save(path)

    def save_bytes(self) -> bytes:
        """Save the workbook to an in-memory byte string."""
        buffer = io.BytesIO()
        self.wb.save(buffer)
        return buffer.getvalue()

    def save_with_postprocess(
        self,
        path: str,
        chart_patches: list[Any] | None = None,
    ) -> None:
        """Save the workbook with OOXML post-processing.

        First saves via openpyxl to get the raw ``.xlsx`` bytes, then
        runs the post-processor to apply Montserrat fonts, pattern
        fills, and asterisk data labels.

        Parameters
        ----------
        path:
            Destination file path.
        chart_patches:
            Optional list of :class:`~autochart.builder.postprocess.ChartPatch`
            objects describing per-chart patches.  If *None* only the
            Montserrat font pass is applied.
        """
        from autochart.builder.postprocess import ChartPatch, postprocess_xlsx

        raw_bytes = self.save_bytes()
        if chart_patches is None:
            chart_patches = []
        processed = postprocess_xlsx(raw_bytes, chart_patches)
        with open(path, "wb") as f:
            f.write(processed)

    # ------------------------------------------------------------------
    # Shared formatting helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _apply_header_style(cell: Cell) -> None:
        """Apply header formatting to *cell*.

        Style: Aptos Narrow 11 pt bold, gray fill (#D9D9D9), thin
        borders on all four sides, horizontally and vertically centred.
        """
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    @staticmethod
    def _apply_data_style(cell: Cell, highlight: bool = False) -> None:
        """Apply data-cell formatting to *cell*.

        Style: Calibri 12 pt, centred.  If *highlight* is ``True`` the
        cell receives a light-blue fill (#DAEEF3).
        """
        cell.font = _DATA_FONT
        cell.alignment = _DATA_ALIGNMENT
        cell.border = _THIN_BORDER
        if highlight:
            cell.fill = _HIGHLIGHT_FILL

    def _create_data_table(
        self,
        ws: Worksheet,
        start_row: int,
        headers: Sequence[str],
        data_rows: Sequence[Sequence[Any]],
        config: ChartConfig | None = None,
    ) -> int:
        """Write a formatted data table and return the next available row.

        Parameters
        ----------
        ws:
            Target worksheet.
        start_row:
            1-based row number where the header row should begin.
        headers:
            Column header labels.
        data_rows:
            An iterable of rows, each an iterable of cell values whose
            length matches *headers*.
        config:
            Optional config override (defaults to ``self.config``).

        Returns
        -------
        int
            The 1-based row number immediately below the last data row.
        """
        if config is None:
            config = self.config

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            self._apply_header_style(cell)

        # Data rows
        current_row = start_row + 1
        for row_data in data_rows:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                self._apply_data_style(cell)
            current_row += 1

        return current_row
