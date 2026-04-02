"""AutoChart input parsers.

Provides :func:`parse_workbook` to detect and parse all INPUT sheets from an
Excel workbook, and :func:`auto_parse` for zero-config usage that auto-detects
configuration metadata from the workbook before parsing.
"""

from pathlib import Path
from typing import Union

import openpyxl

from autochart.config import ChartConfig, ChartSetType, SheetResult
from autochart.parser.base import BaseParser
from autochart.parser.pivoted import PivotedParser
from autochart.parser.sas_output import SASOutputParser


# Registry of available parsers, tried in order
_PARSERS: list[BaseParser] = [
    PivotedParser(),
    SASOutputParser(),
]


def parse_workbook(
    path: Union[str, Path],
    config: ChartConfig,
    sheet_prefix: str = "INPUT",
) -> dict[str, dict]:
    """Open a workbook and parse all INPUT sheets.

    Args:
        path: Path to the Excel workbook (.xlsx).
        config: Chart configuration with disease name, demographics, etc.
        sheet_prefix: Prefix to filter sheets (default: "INPUT").

    Returns:
        A dict mapping sheet names to their parsed results.
        Each parsed result is a dict mapping ChartSetType to the data object(s).

    Example::

        results = parse_workbook("examples.xlsx", config)
        # results["INPUT-1"] == {ChartSetType.A: [ChartSetAData, ...]}
        # results["INPUT-2"] == {ChartSetType.B: [...], ChartSetType.C: ChartSetCData}
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith(sheet_prefix):
            continue

        ws = wb[sheet_name]
        parsed = _parse_sheet(ws, config)
        if parsed:
            results[sheet_name] = parsed

    wb.close()
    return results


def _parse_sheet(worksheet, config: ChartConfig) -> dict:
    """Try each parser on the worksheet and return the first successful result."""
    for parser in _PARSERS:
        if parser.can_parse(worksheet):
            return parser.parse(worksheet, config)
    return {}


def get_all_data_by_type(
    parsed_results: dict[str, dict],
) -> dict[ChartSetType, list]:
    """Aggregate parsed results from all sheets by ChartSetType.

    Args:
        parsed_results: Output from parse_workbook().

    Returns:
        Dict mapping ChartSetType to a flat list of all data objects
        of that type from all sheets.
    """
    by_type: dict[ChartSetType, list] = {}
    for sheet_name, sheet_data in parsed_results.items():
        for chart_type, data in sheet_data.items():
            if chart_type not in by_type:
                by_type[chart_type] = []
            if isinstance(data, list):
                by_type[chart_type].extend(data)
            else:
                by_type[chart_type].append(data)
    return by_type


def auto_parse(
    path: Union[str, Path],
    config_overrides: dict | None = None,
    sheet_prefix: str = "INPUT",
) -> tuple[ChartConfig, dict[ChartSetType, list]]:
    """Auto-detect config from input workbook and parse all sheets.

    Extracts metadata (disease, years, rate, source, geography, demographics)
    from the workbook automatically, merges with any user overrides, then
    parses all INPUT sheets.

    .. note::

        This uses a single aggregated config for all sheets.  For workbooks
        where sheets have different diseases or rate units, use
        :func:`auto_parse_multi` instead.

    Args:
        path: Path to the input .xlsx workbook.
        config_overrides: Optional dict of user-specified config values
            that override auto-detected values. Keys match ChartConfig field names.
        sheet_prefix: Prefix to filter sheets (default: "INPUT").

    Returns:
        Tuple of (auto-built ChartConfig, dict mapping ChartSetType to data lists).
    """
    from autochart.extractor import extract_config, build_config

    extracted = extract_config(path)
    config = build_config(extracted, config_overrides)
    results = parse_workbook(path, config, sheet_prefix)
    by_type = get_all_data_by_type(results)
    return config, by_type


def auto_parse_multi(
    path: Union[str, Path],
    config_overrides: dict | None = None,
    sheet_prefix: str = "INPUT",
) -> list[SheetResult]:
    """Auto-detect per-sheet config and parse each INPUT sheet independently.

    Unlike :func:`auto_parse`, this extracts metadata from each sheet
    separately so that sheets with different diseases or rate units get
    their own :class:`~autochart.config.ChartConfig`.

    When a sheet is missing fields (e.g. no years in its title cells),
    the function falls back to values found in other sheets within the
    same workbook before raising an error.

    Args:
        path: Path to the input .xlsx workbook.
        config_overrides: Optional dict of user-specified config values
            that override auto-detected values on every sheet.
        sheet_prefix: Prefix to filter sheets (default: "INPUT").

    Returns:
        List of :class:`~autochart.config.SheetResult`, one per parsed sheet.
    """
    from autochart.extractor import (
        ExtractedConfig,
        extract_config,
        extract_config_per_sheet,
        build_config,
    )

    per_sheet = extract_config_per_sheet(str(path), sheet_prefix)

    # Build per-disease fallbacks: group sheets by disease name and aggregate
    # values within each group so that e.g. Cancer sheets inherit rate_denom
    # from other Cancer sheets, not from Cerebro sheets.
    disease_groups: dict[str | None, list[str]] = {}
    for sn, ec in per_sheet.items():
        disease_groups.setdefault(ec.disease_name, []).append(sn)

    disease_fallbacks: dict[str | None, ExtractedConfig] = {}
    for disease, sheet_names in disease_groups.items():
        disease_fallbacks[disease] = extract_config(str(path), sheets=sheet_names)

    # Global fallback for truly missing fields
    global_fallback = extract_config(str(path))

    wb = openpyxl.load_workbook(str(path), data_only=True)
    results: list[SheetResult] = []

    for sheet_name, extracted in per_sheet.items():
        # Prefer same-disease fallback, then global fallback
        disease_fb = disease_fallbacks.get(extracted.disease_name, global_fallback)
        merged = ExtractedConfig(
            disease_name=extracted.disease_name or disease_fb.disease_name or global_fallback.disease_name,
            years=extracted.years or disease_fb.years or global_fallback.years,
            rate_unit=extracted.rate_unit or disease_fb.rate_unit or global_fallback.rate_unit,
            rate_denominator=extracted.rate_denominator or disease_fb.rate_denominator or global_fallback.rate_denominator,
            data_source=extracted.data_source or disease_fb.data_source or global_fallback.data_source,
            geography=extracted.geography or disease_fb.geography or global_fallback.geography,
            demographics=extracted.demographics or disease_fb.demographics or global_fallback.demographics,
            reference_group=extracted.reference_group or disease_fb.reference_group or global_fallback.reference_group,
            confidence=extracted.confidence,
        )

        config = build_config(merged, config_overrides)
        ws = wb[sheet_name]
        parsed = _parse_sheet(ws, config)
        if parsed:
            by_type: dict[ChartSetType, list] = {}
            for chart_type, data in parsed.items():
                if isinstance(data, list):
                    by_type[chart_type] = data
                else:
                    by_type[chart_type] = [data]
            results.append(SheetResult(
                sheet_name=sheet_name,
                config=config,
                by_type=by_type,
            ))

    wb.close()
    return results
