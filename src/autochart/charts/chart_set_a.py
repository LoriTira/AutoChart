"""Chart Set A: Race vs Rest of Boston -- WIDE format.

Each race block has a single data row with 9 values across 9 columns,
organized into 3 groups (Boston, Female, Male) with merged headers.
The chart uses a single series with multilevel categories.

Layout per race block (e.g., Asian):
  Row N:   [merged: Boston]  [merged: Female]  [merged: Male]
  Row N+1: Asian | Rest of Boston | Boston Overall  (×3)
  Row N+2: 110.5 | 130.6 | 128.8 | 87.9 | 113.5 | 111.1 | 141.2 | 156.1 | 154.9
  Row N+4: Chart title
  (chart placed below)
"""

from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from autochart.config import ChartConfig, ChartSetAData, ChartSetType
from autochart.text.generator import TextGenerator


def _strip_hash(colour: str) -> str:
    return colour.lstrip("#")


# Style constants matching examples.xlsx
_HEADER_FONT = Font(name="Aptos Narrow", size=11, bold=True, color="000000")
_HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_SUBHEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_LABEL_FONT = Font(name="Aptos Narrow", size=11, color="000000")
_DATA_FONT = Font(name="Calibri", size=12, color="000000")
_DATA_ALIGN = Alignment(horizontal="center", vertical="center")
_HIGHLIGHT_FILL = PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid")

_TITLE_FONT = Font(name="Aptos Narrow", size=11, bold=True, color="000000")
_SHEET_TITLE_FONT = Font(name="Calibri", size=11, bold=True)

_CHART_WIDTH = 15
_CHART_HEIGHT = 7.5
_BLOCK_SPACINGS = [30, 28]  # rows between blocks (3→33=30, 33→61=28)


def build_chart_set_a_sheet(
    ws: Worksheet,
    data_list: list[ChartSetAData],
    config: ChartConfig,
) -> None:
    """Populate *ws* with Chart Set A content in WIDE format."""
    if not data_list:
        return

    text_gen = TextGenerator(config)

    # Column widths (from example OUTPUT-1)
    ws.column_dimensions["A"].width = 20.33
    for letter in "BCDEFGHIJ":
        ws.column_dimensions[letter].width = 15.0

    # Row 1: Sheet title
    cell = ws.cell(row=1, column=1,
                   value="Chart Set A: Race vs Boston overall and Rest of Boston\xa0")
    cell.font = _SHEET_TITLE_FONT
    cell.alignment = Alignment(horizontal="left")

    # Block start rows matching example: 3, 33, 61 (spacing 30, 28)
    block_starts = [3]
    for i in range(1, len(data_list)):
        spacing = _BLOCK_SPACINGS[min(i - 1, len(_BLOCK_SPACINGS) - 1)]
        block_starts.append(block_starts[-1] + spacing)

    for idx, race_data in enumerate(data_list):
        _build_race_block(ws, race_data, config, text_gen, block_starts[idx],
                          title_gap=3 if idx == 0 else 2)


def _build_race_block(
    ws: Worksheet,
    data: ChartSetAData,
    config: ChartConfig,
    text_gen: TextGenerator,
    start_row: int,
    title_gap: int = 3,
) -> None:
    """Build one race block."""
    row = start_row

    # 1. Merged group headers (Boston, Female, Male) spanning 3 cols each
    # Columns: B-D = Boston, E-G = Female, H-J = Male
    groups = ["Boston", "Female", "Male"]
    col_starts = [2, 5, 8]  # B=2, E=5, H=8

    for group_name, col_start in zip(groups, col_starts):
        cell = ws.cell(row=row, column=col_start, value=group_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_start + 2,
        )

    # 2. Sub-headers: [Race, Rest of Boston, Boston Overall] × 3
    sub_row = row + 1
    ws.row_dimensions[sub_row].height = 32.0
    sub_headers = [data.race_name, "Rest of Boston", "Boston Overall"]
    for group_col_start in col_starts:
        for offset, header in enumerate(sub_headers):
            cell = ws.cell(row=sub_row, column=group_col_start + offset, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _SUBHEADER_ALIGN

    # 3. Data row: label + 9 values
    data_row = sub_row + 1
    ws.row_dimensions[data_row].height = 16.0

    # Label in column A
    cell = ws.cell(row=data_row, column=1, value=f"All {config.disease_name}")
    cell.font = _LABEL_FONT

    # 9 data values: [race, rest, overall] × [Boston, Female, Male]
    values = [
        data.boston.group_rate, data.boston.reference_rate, data.boston_overall_rate,
        data.female.group_rate, data.female.reference_rate, data.female_overall_rate,
        data.male.group_rate, data.male.reference_rate, data.male_overall_rate,
    ]

    for i, val in enumerate(values):
        col = 2 + i  # B=2 through J=10
        cell = ws.cell(row=data_row, column=col, value=val)
        cell.font = _DATA_FONT
        cell.alignment = _DATA_ALIGN
        # Highlight first column of each group (race values) — cols B, E, H
        if col in (2, 5, 8):
            cell.fill = _HIGHLIGHT_FILL

    # 4. Chart title (title_gap rows below data)
    title_row = data_row + title_gap
    ws.row_dimensions[title_row].height = 17.0
    title = text_gen.chart_title(ChartSetType.A, race_name=data.race_name)
    cell = ws.cell(row=title_row, column=1, value=title)
    cell.font = _TITLE_FONT

    # 5. Chart: single series from data row, multilevel categories from headers
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 219
    chart.overlap = -27
    chart.legend = None
    chart.title = None

    # Data: single row B{data_row}:J{data_row}
    vals = Reference(ws, min_col=2, max_col=10, min_row=data_row, max_row=data_row)
    chart.add_data(vals, from_rows=True, titles_from_data=False)

    # Categories: multilevel from merged headers + sub-headers
    cats = Reference(ws, min_col=2, max_col=10, min_row=row, max_row=sub_row)
    chart.set_categories(cats)

    # All bars navy (will be post-processed for individual colors)
    series = chart.series[0]
    series.graphicalProperties.solidFill = _strip_hash(config.colors.boston_overall)

    # Data labels
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.showCatName = False
    series.dLbls.showSerName = False
    series.dLbls.dLblPos = "outEnd"

    chart.y_axis.title = f"Rate {config.rate_unit}"
    chart.width = _CHART_WIDTH
    chart.height = _CHART_HEIGHT

    chart_anchor = f"A{title_row + 1}"
    ws.add_chart(chart, chart_anchor)
