"""Chart Set C: Combined race comparison -- single chart, WIDE format.

All races + White + Boston as column headers with a single data row.
The White bar receives diagonal stripe pattern fill via OOXML post-processing.

Layout:
  Row N:   Asian  Black  Latinx  White  Boston  (headers in A-E)
  Row N+1: rate   rate   rate    rate   rate    (values in A-E)
  Row N+3: Chart title
  (chart placed below)
"""

from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from autochart.config import ChartConfig, ChartSetCData, ChartSetType
from autochart.text.generator import TextGenerator


def _strip_hash(colour: str) -> str:
    return colour.lstrip("#")


_HEADER_FONT = Font(name="Aptos Narrow", size=11)
_DATA_FONT = Font(name="Aptos Narrow", size=11)
_TITLE_FONT = Font(name="Aptos Narrow", size=11, bold=True)
_SHEET_TITLE_FONT = Font(name="Calibri", size=11, bold=True)

_CHART_WIDTH = 15
_CHART_HEIGHT = 7.5
_CHART_ROWS = 16


def build_chart_set_c_sheet(
    ws: Worksheet,
    data: ChartSetCData,
    config: ChartConfig,
) -> None:
    """Populate *ws* with Chart Set C content in WIDE format."""
    text_gen = TextGenerator(config)

    _INTRO_FONT = Font(name="Calibri", size=11)

    # Row 1: Sheet title
    cell = ws.cell(row=1, column=1,
                   value="Chart Set C: Combined race comparison chart\xa0")
    cell.font = _SHEET_TITLE_FONT
    cell.alignment = Alignment(horizontal="left")

    # Intro text (matching example OUTPUT-3 rows 2-10)
    intro_lines = [
        "Provide one chart including:\xa0",
        "Bars for all race groups\xa0",
        "Bar for Boston overall\xa0",
        "Bar for White residents\xa0",
        "Confidence intervals (if applicable)\xa0",
        "Chart title\xa0",
        "Axis labels\xa0",
        "Notes/footnotes\xa0",
        "Corresponding descriptive appendix text\xa0",
    ]
    for i, line in enumerate(intro_lines):
        cell = ws.cell(row=2 + i, column=1, value=line)
        cell.font = _INTRO_FONT
        cell.alignment = Alignment(horizontal="left")

    # Build race names + reference group + geography
    race_names = [comp.group_name for comp in data.comparisons]
    all_labels = race_names + [config.reference_group, config.geography]
    num_cols = len(all_labels)

    # Header row (matches example OUTPUT-3 position)
    header_row = 13
    for i, label in enumerate(all_labels):
        col = 1 + i  # A=1, B=2, ...
        cell = ws.cell(row=header_row, column=col, value=label)
        # First column uses Calibri (matching example OUTPUT-3/7 where A13 is Calibri)
        if i == 0:
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.font = _HEADER_FONT

    # Data row
    data_row = header_row + 1
    race_rates = [comp.group_rate for comp in data.comparisons]
    ref_rate = data.comparisons[0].reference_rate
    all_values = race_rates + [ref_rate, data.boston_overall_rate]

    for i, val in enumerate(all_values):
        col = 1 + i
        cell = ws.cell(row=data_row, column=col, value=val)
        cell.font = _DATA_FONT

    # Chart title (2 rows after data)
    title_row = data_row + 2
    ws.row_dimensions[title_row].height = 17.0
    title = text_gen.chart_title(ChartSetType.C)
    cell = ws.cell(row=title_row, column=1, value=title)
    cell.font = _TITLE_FONT

    # Chart: single series from data row
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 219
    chart.overlap = -27
    chart.legend = None
    chart.title = None

    cats = Reference(ws, min_col=1, max_col=num_cols, min_row=header_row)
    vals = Reference(ws, min_col=1, max_col=num_cols, min_row=data_row)
    chart.add_data(vals, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)

    series = chart.series[0]
    series.graphicalProperties.solidFill = _strip_hash(config.colors.boston_overall)

    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.showCatName = False
    series.dLbls.showSerName = False
    series.dLbls.dLblPos = "outEnd"

    chart.y_axis.title = f"Deaths {config.rate_unit}"
    chart.width = _CHART_WIDTH
    chart.height = _CHART_HEIGHT

    chart_anchor = f"A{title_row + 1}"
    ws.add_chart(chart, chart_anchor)

    # Descriptive text
    desc_row = title_row + 1 + 16
    desc_text = text_gen.descriptive_text_set_c(data)
    desc_cell = ws.cell(row=desc_row, column=1, value=desc_text)
    desc_cell.font = Font(name="Calibri", size=10)
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(
        start_row=desc_row, start_column=1,
        end_row=desc_row, end_column=6,
    )

    # Footnote
    footnote_row = desc_row + 2
    footnote_text = text_gen.footnote()
    fn_cell = ws.cell(row=footnote_row, column=1, value=footnote_text)
    fn_cell.font = Font(name="Calibri", size=8, color="595959")
    fn_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(
        start_row=footnote_row, start_column=1,
        end_row=footnote_row + 1, end_column=6,
    )
