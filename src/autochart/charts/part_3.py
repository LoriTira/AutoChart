"""Part 3: Sex- and Race-Stratified Charts -- single chart, WIDE format.

Merged Female/Male headers spanning 5 columns each, with race sub-headers
underneath and a single data row of 10 values.

Layout:
  Row N:   [merged: Female (B:F)]  [merged: Male (G:K)]
  Row N+1: Asian | Black | Latinx | White | Boston  × 2
  Row N+2: 10 values (B through K)
  Row N+4: Chart title
  (chart placed below)
"""

from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from autochart.config import ChartConfig, ChartSetType, Part3Data
from autochart.text.generator import TextGenerator


def _strip_hash(colour: str) -> str:
    return colour.lstrip("#")


_HEADER_FONT = Font(name="Aptos Narrow", size=11)
_DATA_FONT = Font(name="Aptos Narrow", size=11)
_TITLE_FONT = Font(name="Aptos Narrow", size=11, bold=True)
_SHEET_TITLE_FONT = Font(name="Calibri", size=11, bold=True)
_HIGHLIGHT_FILL = PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid")

_CHART_WIDTH = 18
_CHART_HEIGHT = 7.5
_CHART_ROWS = 16


def build_part_3_sheet(
    ws: Worksheet,
    data: Part3Data,
    config: ChartConfig,
) -> None:
    """Populate *ws* with Part 3 content matching example format."""
    text_gen = TextGenerator(config)

    # Column widths (from example OUTPUT-4/8)
    ws.column_dimensions["A"].width = 12.83
    ws.column_dimensions["B"].width = 12.0
    ws.column_dimensions["D"].width = 12.5

    # Row 1: Sheet title
    cell = ws.cell(row=1, column=1,
                   value="Part 3: Sex- and Race-Stratified Charts")
    cell.font = _SHEET_TITLE_FONT
    cell.alignment = Alignment(horizontal="left")

    # Race names from data
    race_names = [comp.group_name for comp in data.female_comparisons]
    sub_labels = race_names + [config.reference_group, config.geography]
    n_sub = len(sub_labels)  # typically 5

    # Merged group headers (row 3)
    header_row = 3
    # Female: B3:F3
    cell = ws.cell(row=header_row, column=2, value="Female")
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(
        start_row=header_row, start_column=2,
        end_row=header_row, end_column=1 + n_sub,
    )
    # Male: G3:K3
    male_start = 2 + n_sub
    cell = ws.cell(row=header_row, column=male_start, value="Male")
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(
        start_row=header_row, start_column=male_start,
        end_row=header_row, end_column=male_start + n_sub - 1,
    )

    # Sub-headers (row 4): race names × 2
    sub_row = header_row + 1
    for i, label in enumerate(sub_labels):
        # Female group
        cell = ws.cell(row=sub_row, column=2 + i, value=label)
        cell.font = _HEADER_FONT
        # Male group
        cell = ws.cell(row=sub_row, column=male_start + i, value=label)
        cell.font = _HEADER_FONT

    # Data row (row 5): 10 values
    data_row = sub_row + 1

    female_rates = [comp.group_rate for comp in data.female_comparisons]
    female_ref_rate = data.female_comparisons[0].reference_rate
    female_values = female_rates + [female_ref_rate, data.female_boston_rate]

    male_rates = [comp.group_rate for comp in data.male_comparisons]
    male_ref_rate = data.male_comparisons[0].reference_rate
    male_values = male_rates + [male_ref_rate, data.male_boston_rate]

    all_values = female_values + male_values

    for i, val in enumerate(all_values):
        col = 2 + i
        cell = ws.cell(row=data_row, column=col, value=val)
        cell.font = _DATA_FONT
        # Highlight significant race values with blue fill
        # (significant comparisons get highlighted)
        if i < len(data.female_comparisons):
            comp = data.female_comparisons[i]
            if comp.p_value is not None and comp.p_value < config.significance_threshold:
                cell.fill = _HIGHLIGHT_FILL
        elif i >= n_sub and i < n_sub + len(data.male_comparisons):
            comp = data.male_comparisons[i - n_sub]
            if comp.p_value is not None and comp.p_value < config.significance_threshold:
                cell.fill = _HIGHLIGHT_FILL

    # Chart title (2 rows after data)
    title_row = data_row + 3
    ws.row_dimensions[title_row].height = 17.0
    title = text_gen.chart_title(ChartSetType.PART_3)
    cell = ws.cell(row=title_row, column=2, value=title)
    cell.font = _TITLE_FONT

    # Chart: single series, 10 data points, multilevel categories
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 219
    chart.overlap = -27
    chart.legend = None
    chart.title = None

    total_cols = 2 * n_sub
    vals = Reference(ws, min_col=2, max_col=1 + total_cols, min_row=data_row)
    chart.add_data(vals, from_rows=True, titles_from_data=False)

    # Multilevel categories from merged headers + sub-headers
    cats = Reference(ws, min_col=2, max_col=1 + total_cols,
                     min_row=header_row, max_row=sub_row)
    chart.set_categories(cats)

    series = chart.series[0]
    series.graphicalProperties.solidFill = _strip_hash(config.colors.boston_overall)

    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.showCatName = False
    series.dLbls.showSerName = False
    series.dLbls.dLblPos = "outEnd"

    chart.y_axis.title = f"Rate {config.rate_unit}"
    chart.width = _CHART_WIDTH
    chart.height = _CHART_HEIGHT

    chart_anchor = f"A{title_row + 1}"
    ws.add_chart(chart, chart_anchor)
