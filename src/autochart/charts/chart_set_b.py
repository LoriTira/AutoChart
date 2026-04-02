"""Chart Set B: Race vs White (reference group) -- one chart per race.

Each chart has 3 bars: [Race, White, Boston], all navy.
The White bar receives a diagonal stripe pattern fill via OOXML post-processing.

Layout per race block:
  Row N:   [Race] [White] [Boston]   (headers in B/C/D, Aptos Narrow 11pt)
  Row N+1: rate    rate    rate      (values in B/C/D, Aptos Narrow 11pt)
  Row N+3: Chart title
  (chart placed below)
"""

from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from autochart.config import ChartConfig, ChartSetBData, ChartSetType
from autochart.text.generator import TextGenerator


def _strip_hash(colour: str) -> str:
    return colour.lstrip("#")


_HEADER_FONT = Font(name="Aptos Narrow", size=11)
_DATA_FONT = Font(name="Aptos Narrow", size=11)
_TITLE_FONT = Font(name="Aptos Narrow", size=11, bold=True)
_SHEET_TITLE_FONT = Font(name="Calibri", size=11, bold=True)

_CHART_WIDTH = 15
_CHART_HEIGHT = 7.5
_CHART_ROWS = 16


def build_chart_set_b_sheet(
    ws: Worksheet,
    data_list: list[ChartSetBData],
    config: ChartConfig,
) -> None:
    """Populate *ws* with Chart Set B content matching example format."""
    if not data_list:
        return

    text_gen = TextGenerator(config)

    # Column widths (from example OUTPUT-6)
    ws.column_dimensions["A"].width = 12.66
    ws.column_dimensions["B"].width = 11.16
    ws.column_dimensions["C"].width = 11.5
    ws.column_dimensions["D"].width = 12.0

    # Row 1: Sheet title
    cell = ws.cell(row=1, column=1,
                   value="Chart Set B: Race vs White residents (reference group)")
    cell.font = _SHEET_TITLE_FONT
    cell.alignment = Alignment(horizontal="left")

    current_row = 5  # first block starts at row 5

    for race_data in data_list:
        current_row = _build_race_block(ws, race_data, config, text_gen, current_row)
        current_row += 2


def _build_race_block(
    ws: Worksheet,
    data: ChartSetBData,
    config: ChartConfig,
    text_gen: TextGenerator,
    start_row: int,
) -> int:
    """Build one race block and return the next available row."""
    row = start_row

    # 1. Headers: Race, White, Boston (cols B, C, D)
    headers = [data.race_name, config.reference_group, config.geography]
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=2 + i, value=header)
        cell.font = _HEADER_FONT

    # 2. Data row (cols B, C, D)
    data_row = row + 1
    values = [
        data.comparison.group_rate,
        data.comparison.reference_rate,
        data.boston_overall_rate,
    ]
    for i, val in enumerate(values):
        cell = ws.cell(row=data_row, column=2 + i, value=val)
        cell.font = _DATA_FONT

    # 3. Chart title (2 rows after data)
    title_row = data_row + 2
    ws.row_dimensions[title_row].height = 17.0
    title = text_gen.chart_title(ChartSetType.B, race_name=data.race_name)
    cell = ws.cell(row=title_row, column=1, value=title)
    cell.font = _TITLE_FONT

    # 4. Chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 219
    chart.overlap = -27
    chart.legend = None
    chart.title = None

    cats = Reference(ws, min_col=2, max_col=4, min_row=row)
    vals = Reference(ws, min_col=2, max_col=4, min_row=data_row)
    chart.add_data(vals, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)

    series = chart.series[0]
    series.graphicalProperties.solidFill = _strip_hash(config.colors.boston_overall)

    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.showCatName = False
    series.dLbls.showSerName = False
    series.dLbls.dLblPos = "outEnd"

    chart.y_axis.title = f"Rate {config.rate_unit}"
    chart.width = _CHART_WIDTH
    chart.height = _CHART_HEIGHT

    chart_anchor = f"A{title_row + 1}"
    ws.add_chart(chart, chart_anchor)

    return title_row + 1 + _CHART_ROWS
