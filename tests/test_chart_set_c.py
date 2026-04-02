"""Tests for Chart Set C: Combined race comparison -- WIDE format."""

from __future__ import annotations

import io

import openpyxl
import pytest

from autochart.builder.workbook import WorkbookBuilder
from autochart.charts.chart_set_c import build_chart_set_c_sheet
from autochart.config import ChartConfig, ChartSetCData, RateComparison


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def config() -> ChartConfig:
    return ChartConfig(
        disease_name="Cancer Mortality",
        rate_unit="per 100,000 residents",
        rate_denominator=100_000,
        data_source="DATA SOURCE: Boston resident deaths, Massachusetts Department of Public Health",
        years="2017-2023",
    )


def _make_comparison(
    group_name: str,
    group_rate: float,
    reference_name: str,
    reference_rate: float,
    p_value: float | None = None,
) -> RateComparison:
    return RateComparison(
        group_name=group_name,
        group_rate=group_rate,
        reference_name=reference_name,
        reference_rate=reference_rate,
        p_value=p_value,
    )


@pytest.fixture()
def chart_set_c_data() -> ChartSetCData:
    return ChartSetCData(
        comparisons=[
            _make_comparison("Asian", 110.5, "White", 125.3, p_value=0.03),
            _make_comparison("Black", 153.0, "White", 125.3, p_value=0.0001),
            _make_comparison("Latinx", 99.5, "White", 125.3, p_value=0.0001),
        ],
        boston_overall_rate=128.8,
    )


@pytest.fixture()
def builder(config: ChartConfig) -> WorkbookBuilder:
    return WorkbookBuilder(config)


# ---------------------------------------------------------------------------
# build_chart_set_c_sheet -- direct function tests
# ---------------------------------------------------------------------------

class TestBuildChartSetCSheet:
    """Test the build_chart_set_c_sheet function directly."""

    def test_sheet_title(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        assert ws.cell(row=1, column=1).value == "Chart Set C: Combined race comparison chart\xa0"

    def test_header_row_at_13(self, config, chart_set_c_data):
        """WIDE format: headers at row 13 across 5 columns (A-E)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        assert ws.cell(row=13, column=1).value == "Asian"
        assert ws.cell(row=13, column=2).value == "Black"
        assert ws.cell(row=13, column=3).value == "Latinx"
        assert ws.cell(row=13, column=4).value == "White"
        assert ws.cell(row=13, column=5).value == "Boston"

    def test_data_row_at_14(self, config, chart_set_c_data):
        """WIDE format: data row at row 14 with 5 values."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        assert ws.cell(row=14, column=1).value == 110.5  # Asian
        assert ws.cell(row=14, column=2).value == 153.0  # Black
        assert ws.cell(row=14, column=3).value == 99.5   # Latinx
        assert ws.cell(row=14, column=4).value == 125.3  # White
        assert ws.cell(row=14, column=5).value == 128.8  # Boston Overall

    def test_chart_title_below_data(self, config, chart_set_c_data):
        """Chart title is 2 rows below data (row 16)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        title = str(ws.cell(row=16, column=1).value)
        assert "by Race" in title or "2017-2023" in title

    def test_chart_created(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        assert len(ws._charts) == 1

    def test_chart_has_one_series(self, config, chart_set_c_data):
        """Single series from row data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        chart = ws._charts[0]
        assert len(chart.series) == 1

    def test_chart_type_is_col(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        chart = ws._charts[0]
        assert chart.type == "col"
        assert chart.grouping == "clustered"

    def test_chart_gap_and_overlap(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        chart = ws._charts[0]
        assert chart.gapWidth == 219
        assert chart.overlap == -27

    def test_data_labels_enabled(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        chart = ws._charts[0]
        series = chart.series[0]
        assert series.dLbls is not None
        assert series.dLbls.showVal is True

    def test_descriptive_text_present(self, config, chart_set_c_data):
        """New format has no descriptive text."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        found_descriptive = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "age-adjusted" in str(row[0]).lower():
                found_descriptive = True
                break
        assert found_descriptive, "Descriptive text should be present"

    def test_footnote_present(self, config, chart_set_c_data):
        """New format has no footnotes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        found_footnote = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "DATA SOURCE" in str(row[0]):
                found_footnote = True
                break
        assert found_footnote, "Footnote should be present"

    def test_chart_has_no_legend(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        chart = ws._charts[0]
        assert chart.legend is None

    def test_y_axis_has_title(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        chart = ws._charts[0]
        assert chart.y_axis.title is not None
        assert "per 100,000" in str(chart.y_axis.title)

    def test_series_fill_is_set(self, config, chart_set_c_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_c_sheet(ws, chart_set_c_data, config)
        chart = ws._charts[0]
        assert chart.series[0].graphicalProperties.solidFill is not None


# ---------------------------------------------------------------------------
# WorkbookBuilder.add_chart_set_c integration tests
# ---------------------------------------------------------------------------

class TestAddChartSetCIntegration:
    """Test the add_chart_set_c method via WorkbookBuilder."""

    def test_creates_output_3_sheet(self, builder, chart_set_c_data):
        builder.add_chart_set_c(chart_set_c_data)
        assert "OUTPUT-3" in builder.wb.sheetnames

    def test_sheet_has_one_chart(self, builder, chart_set_c_data):
        builder.add_chart_set_c(chart_set_c_data)
        ws = builder.wb["OUTPUT-3"]
        assert len(ws._charts) == 1

    def test_save_bytes_produces_valid_xlsx(self, builder, chart_set_c_data):
        builder.add_chart_set_c(chart_set_c_data)
        data = builder.save_bytes()
        assert data[:2] == b"PK"
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        assert "OUTPUT-3" in loaded.sheetnames

    def test_save_and_reload_preserves_data(self, builder, chart_set_c_data):
        builder.add_chart_set_c(chart_set_c_data)
        data = builder.save_bytes()
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        ws = loaded["OUTPUT-3"]
        # WIDE format: row 14 has data values
        assert ws.cell(row=14, column=1).value == 110.5
        assert ws.cell(row=14, column=2).value == 153.0

    def test_multiple_calls_create_unique_sheets(self, builder, chart_set_c_data):
        builder.add_chart_set_c(chart_set_c_data)
        builder.add_chart_set_c(chart_set_c_data)
        assert "OUTPUT-3" in builder.wb.sheetnames
        assert "OUTPUT-3_2" in builder.wb.sheetnames
