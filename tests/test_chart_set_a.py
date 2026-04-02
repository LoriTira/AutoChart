"""Tests for Chart Set A: Race vs Rest of Boston -- WIDE format."""

from __future__ import annotations

import io

import openpyxl
import pytest

from autochart.builder.workbook import WorkbookBuilder
from autochart.charts.chart_set_a import build_chart_set_a_sheet
from autochart.config import ChartConfig, ChartSetAData, RateComparison


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def config() -> ChartConfig:
    return ChartConfig(
        disease_name="Cancer Mortality",
        rate_unit="per 100,000 residents",
        rate_denominator=100_000,
        data_source="DATA SOURCE: Boston resident deaths, Massachusetts Department of Public Health",
        years="2017-2023",
    )


def _make_comparison(
    group_name: str,
    group_rate: float,
    reference_name: str,
    reference_rate: float,
    p_value: float | None = None,
) -> RateComparison:
    return RateComparison(
        group_name=group_name,
        group_rate=group_rate,
        reference_name=reference_name,
        reference_rate=reference_rate,
        p_value=p_value,
    )


@pytest.fixture()
def asian_data() -> ChartSetAData:
    return ChartSetAData(
        race_name="Asian",
        boston=_make_comparison("Asian", 110.5, "Rest of Boston", 130.6, p_value=0.02),
        female=_make_comparison("Asian", 87.9, "Rest of Boston", 113.5, p_value=0.01),
        male=_make_comparison("Asian", 141.1, "Rest of Boston", 152.9, p_value=0.35),
        boston_overall_rate=128.8,
        female_overall_rate=111.5,
        male_overall_rate=150.8,
    )


@pytest.fixture()
def black_data() -> ChartSetAData:
    return ChartSetAData(
        race_name="Black",
        boston=_make_comparison("Black", 153.0, "Rest of Boston", 122.2, p_value=0.0001),
        female=_make_comparison("Black", 127.7, "Rest of Boston", 105.4, p_value=0.005),
        male=_make_comparison("Black", 188.5, "Rest of Boston", 144.5, p_value=0.0001),
        boston_overall_rate=128.8,
        female_overall_rate=111.5,
        male_overall_rate=150.8,
    )


@pytest.fixture()
def latinx_data() -> ChartSetAData:
    return ChartSetAData(
        race_name="Latinx",
        boston=_make_comparison("Latinx", 99.5, "Rest of Boston", 132.9, p_value=0.0001),
        female=_make_comparison("Latinx", 88.5, "Rest of Boston", 114.6, p_value=0.005),
        male=_make_comparison("Latinx", 113.9, "Rest of Boston", 156.2, p_value=0.0001),
        boston_overall_rate=128.8,
        female_overall_rate=111.5,
        male_overall_rate=150.8,
    )


@pytest.fixture()
def data_list(asian_data, black_data, latinx_data) -> list[ChartSetAData]:
    return [asian_data, black_data, latinx_data]


@pytest.fixture()
def builder(config: ChartConfig) -> WorkbookBuilder:
    return WorkbookBuilder(config)


# ---------------------------------------------------------------------------
# build_chart_set_a_sheet -- direct function tests
# ---------------------------------------------------------------------------

class TestBuildChartSetASheet:
    """Test the build_chart_set_a_sheet function directly."""

    def test_no_crash_empty_list(self, config):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [], config)
        # Should return without error; no data written
        assert ws.cell(row=1, column=1).value is None

    def test_sheet_title(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        assert ws.cell(row=1, column=1).value == "Chart Set A: Race vs Boston overall and Rest of Boston"

    def test_merged_group_headers(self, config, asian_data):
        """First race block starts at row 3 with merged headers: Boston, Female, Male."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        # Merged headers at row 3
        assert ws.cell(row=3, column=2).value == "Boston"   # B3
        assert ws.cell(row=3, column=5).value == "Female"   # E3
        assert ws.cell(row=3, column=8).value == "Male"     # H3

    def test_sub_headers(self, config, asian_data):
        """Sub-headers at row 4: [Race, Rest of Boston, Boston Overall] x3."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        # Sub-headers at row 4 for all 3 groups
        assert ws.cell(row=4, column=2).value == "Asian"
        assert ws.cell(row=4, column=3).value == "Rest of Boston"
        assert ws.cell(row=4, column=4).value == "Boston Overall"
        assert ws.cell(row=4, column=5).value == "Asian"
        assert ws.cell(row=4, column=6).value == "Rest of Boston"
        assert ws.cell(row=4, column=7).value == "Boston Overall"
        assert ws.cell(row=4, column=8).value == "Asian"
        assert ws.cell(row=4, column=9).value == "Rest of Boston"
        assert ws.cell(row=4, column=10).value == "Boston Overall"

    def test_data_row_label(self, config, asian_data):
        """Data row has 'All {Disease}' label in col A at row 5."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        assert ws.cell(row=5, column=1).value == "All Cancer Mortality"

    def test_data_row_values(self, config, asian_data):
        """Data row 5: 9 values across B-J in WIDE format."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        # Boston group: race, rest, overall
        assert ws.cell(row=5, column=2).value == 110.5   # Asian (Boston)
        assert ws.cell(row=5, column=3).value == 130.6   # Rest of Boston
        assert ws.cell(row=5, column=4).value == 128.8   # Boston Overall
        # Female group
        assert ws.cell(row=5, column=5).value == 87.9    # Asian (Female)
        assert ws.cell(row=5, column=6).value == 113.5   # Rest of Boston
        assert ws.cell(row=5, column=7).value == 111.5   # Female Overall
        # Male group
        assert ws.cell(row=5, column=8).value == 141.1   # Asian (Male)
        assert ws.cell(row=5, column=9).value == 152.9   # Rest of Boston
        assert ws.cell(row=5, column=10).value == 150.8  # Male Overall

    def test_chart_title_below_data(self, config, asian_data):
        """Chart title is 2 rows below data row (row 7)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        title_val = str(ws.cell(row=7, column=1).value)
        assert "Asian" in title_val
        assert "2017-2023" in title_val

    def test_single_race_chart_created(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        assert len(ws._charts) == 1

    def test_three_races_three_charts(self, config, data_list):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, data_list, config)
        assert len(ws._charts) == 3

    def test_chart_has_single_series(self, config, asian_data):
        """WIDE format uses a single series with 9 data points."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        chart = ws._charts[0]
        assert len(chart.series) == 1

    def test_chart_type_is_col(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        chart = ws._charts[0]
        assert chart.type == "col"
        assert chart.grouping == "clustered"

    def test_chart_gap_and_overlap(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        chart = ws._charts[0]
        assert chart.gapWidth == 219
        assert chart.overlap == -27

    def test_data_labels_enabled(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        chart = ws._charts[0]
        series = chart.series[0]
        assert series.dLbls is not None
        assert series.dLbls.showVal is True

    def test_no_descriptive_text(self, config, asian_data):
        """New format has no descriptive text or footnotes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        found_descriptive = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "age-adjusted" in str(row[0]).lower():
                found_descriptive = True
                break
        assert not found_descriptive, "Descriptive text should not be present in new format"

    def test_no_footnote(self, config, asian_data):
        """New format has no footnotes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        found_footnote = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "DATA SOURCE" in str(row[0]):
                found_footnote = True
                break
        assert not found_footnote, "Footnote should not be present in new format"

    def test_header_cells_have_gray_fill(self, config, asian_data):
        """Fill is E8E8E8 (was D9D9D9)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        header_cell = ws.cell(row=3, column=2)
        assert header_cell.fill.start_color.rgb == "00E8E8E8"

    def test_race_column_highlighted(self, config, asian_data):
        """Highlight is CAEDFB (was DAEEF3). Applied to race value cols B, E, H."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        for col in (2, 5, 8):
            cell = ws.cell(row=5, column=col)
            assert cell.fill.start_color.rgb == "00CAEDFB", (
                f"Col {col} row 5 should have highlight fill CAEDFB"
            )


# ---------------------------------------------------------------------------
# WorkbookBuilder.add_chart_set_a integration tests
# ---------------------------------------------------------------------------

class TestAddChartSetAIntegration:
    """Test the add_chart_set_a method via WorkbookBuilder."""

    def test_empty_data_list_no_sheet(self, builder):
        builder.add_chart_set_a([])
        assert "OUTPUT-1" not in builder.wb.sheetnames

    def test_creates_output_1_sheet(self, builder, data_list):
        builder.add_chart_set_a(data_list)
        assert "OUTPUT-1" in builder.wb.sheetnames

    def test_sheet_has_three_charts(self, builder, data_list):
        builder.add_chart_set_a(data_list)
        ws = builder.wb["OUTPUT-1"]
        assert len(ws._charts) == 3

    def test_save_bytes_produces_valid_xlsx(self, builder, data_list):
        builder.add_chart_set_a(data_list)
        data = builder.save_bytes()
        assert data[:2] == b"PK"
        # Loadable by openpyxl
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        assert "OUTPUT-1" in loaded.sheetnames

    def test_save_and_reload_preserves_data(self, builder, asian_data):
        builder.add_chart_set_a([asian_data])
        data = builder.save_bytes()
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        ws = loaded["OUTPUT-1"]
        # Check WIDE data row values survive the round-trip (row 5, cols B-J)
        assert ws.cell(row=5, column=2).value == 110.5
        assert ws.cell(row=5, column=5).value == 87.9
        assert ws.cell(row=5, column=8).value == 141.1

    def test_all_three_race_blocks_have_data(self, builder, data_list):
        builder.add_chart_set_a(data_list)
        ws = builder.wb["OUTPUT-1"]
        # Verify each race name appears somewhere in the sheet
        all_values = []
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=10, values_only=True):
            all_values.extend([str(v) for v in row if v is not None])
        joined = " ".join(all_values)
        assert "Asian" in joined
        assert "Black" in joined
        assert "Latinx" in joined


# ---------------------------------------------------------------------------
# Chart colour tests
# ---------------------------------------------------------------------------

class TestChartColours:
    """Verify series fill colours are set correctly."""

    def test_single_series_has_fill(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        chart = ws._charts[0]
        assert len(chart.series) == 1
        fill = chart.series[0].graphicalProperties.solidFill
        assert fill is not None

    def test_chart_has_no_legend(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        chart = ws._charts[0]
        assert chart.legend is None

    def test_y_axis_has_title(self, config, asian_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_a_sheet(ws, [asian_data], config)
        chart = ws._charts[0]
        assert chart.y_axis.title is not None
        assert "per 100,000" in str(chart.y_axis.title)
