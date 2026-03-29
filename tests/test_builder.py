"""Tests for autochart.builder.workbook -- WorkbookBuilder."""

from __future__ import annotations

import io

import openpyxl
import pytest

from autochart.builder.workbook import WorkbookBuilder
from autochart.config import ChartConfig


@pytest.fixture()
def config() -> ChartConfig:
    return ChartConfig(
        disease_name="Cancer Mortality",
        rate_unit="per 100,000 residents",
        rate_denominator=100_000,
        data_source="DATA SOURCE: Test",
        years="2017-2023",
    )


@pytest.fixture()
def builder(config: ChartConfig) -> WorkbookBuilder:
    return WorkbookBuilder(config)


# -----------------------------------------------------------------------
# Instantiation
# -----------------------------------------------------------------------

class TestWorkbookBuilderInit:
    def test_can_instantiate(self, config):
        wb = WorkbookBuilder(config)
        assert wb is not None

    def test_config_stored(self, config):
        wb = WorkbookBuilder(config)
        assert wb.config is config

    def test_workbook_created(self, builder):
        assert builder.wb is not None
        assert isinstance(builder.wb, openpyxl.Workbook)

    def test_default_sheet_removed(self, builder):
        assert len(builder.wb.sheetnames) == 0

    def test_save_bytes_returns_valid_xlsx(self, builder):
        # Create at least one sheet so save succeeds
        builder.wb.create_sheet("Test")
        data = builder.save_bytes()
        assert isinstance(data, bytes)
        assert len(data) > 0
        # Should be valid xlsx (ZIP starts with PK)
        assert data[:2] == b"PK"

    def test_save_bytes_loadable(self, builder):
        builder.wb.create_sheet("Test")
        data = builder.save_bytes()
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        assert "Test" in loaded.sheetnames


# -----------------------------------------------------------------------
# Data table creation
# -----------------------------------------------------------------------

class TestCreateDataTable:
    def test_headers_written(self, builder):
        ws = builder.wb.create_sheet("DT")
        headers = ["Race", "Rate", "p-value"]
        builder._create_data_table(ws, 1, headers, [])
        for col, header in enumerate(headers, start=1):
            assert ws.cell(row=1, column=col).value == header

    def test_data_rows_written(self, builder):
        ws = builder.wb.create_sheet("DT")
        headers = ["Race", "Rate"]
        data = [["Asian", 12.3], ["Black", 15.6]]
        builder._create_data_table(ws, 1, headers, data)
        assert ws.cell(row=2, column=1).value == "Asian"
        assert ws.cell(row=2, column=2).value == 12.3
        assert ws.cell(row=3, column=1).value == "Black"
        assert ws.cell(row=3, column=2).value == 15.6

    def test_returns_next_row(self, builder):
        ws = builder.wb.create_sheet("DT")
        headers = ["A", "B"]
        data = [["x", 1], ["y", 2], ["z", 3]]
        next_row = builder._create_data_table(ws, 5, headers, data)
        # Header at row 5, data at rows 6-8, next should be 9
        assert next_row == 9

    def test_empty_data(self, builder):
        ws = builder.wb.create_sheet("DT")
        next_row = builder._create_data_table(ws, 1, ["H"], [])
        assert next_row == 2


# -----------------------------------------------------------------------
# Cell styles
# -----------------------------------------------------------------------

class TestHeaderStyle:
    def test_font_name(self, builder):
        ws = builder.wb.create_sheet("Style")
        cell = ws.cell(row=1, column=1, value="Header")
        builder._apply_header_style(cell)
        assert cell.font.name == "Aptos Narrow"

    def test_font_bold(self, builder):
        ws = builder.wb.create_sheet("Style")
        cell = ws.cell(row=1, column=1, value="Header")
        builder._apply_header_style(cell)
        assert cell.font.bold is True

    def test_font_size(self, builder):
        ws = builder.wb.create_sheet("Style")
        cell = ws.cell(row=1, column=1, value="Header")
        builder._apply_header_style(cell)
        assert cell.font.size == 11

    def test_fill_is_gray(self, builder):
        ws = builder.wb.create_sheet("Style")
        cell = ws.cell(row=1, column=1, value="Header")
        builder._apply_header_style(cell)
        assert cell.fill.start_color.rgb == "00D9D9D9"

    def test_alignment_centered(self, builder):
        ws = builder.wb.create_sheet("Style")
        cell = ws.cell(row=1, column=1, value="Header")
        builder._apply_header_style(cell)
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"

    def test_borders_thin(self, builder):
        ws = builder.wb.create_sheet("Style")
        cell = ws.cell(row=1, column=1, value="Header")
        builder._apply_header_style(cell)
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"


class TestDataStyle:
    def test_font_name(self, builder):
        ws = builder.wb.create_sheet("DStyle")
        cell = ws.cell(row=1, column=1, value=42)
        builder._apply_data_style(cell)
        assert cell.font.name == "Calibri"

    def test_font_size(self, builder):
        ws = builder.wb.create_sheet("DStyle")
        cell = ws.cell(row=1, column=1, value=42)
        builder._apply_data_style(cell)
        assert cell.font.size == 12

    def test_no_highlight_by_default(self, builder):
        ws = builder.wb.create_sheet("DStyle")
        cell = ws.cell(row=1, column=1, value=42)
        builder._apply_data_style(cell)
        # Should not have the highlight fill
        assert cell.fill.fill_type is None or cell.fill.start_color.rgb != "00DAEEF3"

    def test_highlight_fill(self, builder):
        ws = builder.wb.create_sheet("DStyle")
        cell = ws.cell(row=1, column=1, value=42)
        builder._apply_data_style(cell, highlight=True)
        assert cell.fill.start_color.rgb == "00DAEEF3"

    def test_alignment_centered(self, builder):
        ws = builder.wb.create_sheet("DStyle")
        cell = ws.cell(row=1, column=1, value=42)
        builder._apply_data_style(cell)
        assert cell.alignment.horizontal == "center"

    def test_borders_thin(self, builder):
        ws = builder.wb.create_sheet("DStyle")
        cell = ws.cell(row=1, column=1, value=42)
        builder._apply_data_style(cell)
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"


# -----------------------------------------------------------------------
# Stub methods exist
# -----------------------------------------------------------------------

class TestStubMethods:
    def test_add_chart_set_a_exists(self, builder):
        assert hasattr(builder, "add_chart_set_a")

    def test_add_chart_set_b_exists(self, builder):
        assert hasattr(builder, "add_chart_set_b")

    def test_add_chart_set_c_exists(self, builder):
        assert hasattr(builder, "add_chart_set_c")

    def test_add_part_3_exists(self, builder):
        assert hasattr(builder, "add_part_3")

    def test_stubs_are_callable(self, builder):
        # Should not raise
        builder.add_chart_set_a([])
        builder.add_chart_set_b([])
