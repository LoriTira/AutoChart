"""Tests for Chart Set B: Race vs White (reference group) -- WIDE format."""

from __future__ import annotations

import io

import openpyxl
import pytest

from autochart.builder.workbook import WorkbookBuilder
from autochart.charts.chart_set_b import build_chart_set_b_sheet
from autochart.config import ChartConfig, ChartSetBData, RateComparison


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def config() -> ChartConfig:
    return ChartConfig(
        disease_name="Cancer Mortality",
        rate_unit="per 100,000 residents",
        rate_denominator=100_000,
        data_source="DATA SOURCE: Boston resident deaths, Massachusetts Department of Public Health",
        years="2017-2023",
    )


def _make_comparison(
    group_name: str,
    group_rate: float,
    reference_name: str,
    reference_rate: float,
    p_value: float | None = None,
) -> RateComparison:
    return RateComparison(
        group_name=group_name,
        group_rate=group_rate,
        reference_name=reference_name,
        reference_rate=reference_rate,
        p_value=p_value,
    )


@pytest.fixture()
def asian_b_data() -> ChartSetBData:
    return ChartSetBData(
        race_name="Asian",
        comparison=_make_comparison("Asian", 110.5, "White", 125.3, p_value=0.03),
        boston_overall_rate=128.8,
    )


@pytest.fixture()
def black_b_data() -> ChartSetBData:
    return ChartSetBData(
        race_name="Black",
        comparison=_make_comparison("Black", 153.0, "White", 125.3, p_value=0.0001),
        boston_overall_rate=128.8,
    )


@pytest.fixture()
def latinx_b_data() -> ChartSetBData:
    return ChartSetBData(
        race_name="Latinx",
        comparison=_make_comparison("Latinx", 99.5, "White", 125.3, p_value=0.0001),
        boston_overall_rate=128.8,
    )


@pytest.fixture()
def data_list(asian_b_data, black_b_data, latinx_b_data) -> list[ChartSetBData]:
    return [asian_b_data, black_b_data, latinx_b_data]


@pytest.fixture()
def builder(config: ChartConfig) -> WorkbookBuilder:
    return WorkbookBuilder(config)


# ---------------------------------------------------------------------------
# build_chart_set_b_sheet -- direct function tests
# ---------------------------------------------------------------------------

class TestBuildChartSetBSheet:
    """Test the build_chart_set_b_sheet function directly."""

    def test_no_crash_empty_list(self, config):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [], config)
        # Should return without error; no data written
        assert ws.cell(row=1, column=1).value is None

    def test_sheet_title(self, config, asian_b_data):
        """Row 1 has the sheet title."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        assert ws.cell(row=1, column=1).value == "Chart Set B: Race vs White residents (reference group)\xa0"

    def test_first_block_starts_at_row_5(self, config, asian_b_data):
        """First block headers start at row 5."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        # Headers in B/C/D at row 5
        assert ws.cell(row=5, column=2).value == "Asian"
        assert ws.cell(row=5, column=3).value == "White"
        assert ws.cell(row=5, column=4).value == "Boston"

    def test_data_row_values(self, config, asian_b_data):
        """Data in B/C/D at row 6 (row after headers)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        assert ws.cell(row=6, column=2).value == 110.5
        assert ws.cell(row=6, column=3).value == 125.3
        assert ws.cell(row=6, column=4).value == 128.8

    def test_no_aar_label(self, config, asian_b_data):
        """No 'AAR' label in column A data row."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        assert ws.cell(row=6, column=1).value is None

    def test_no_gray_fill_on_headers(self, config, asian_b_data):
        """Headers have no fill (Aptos Narrow 11pt, no bold, no fill)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        header_cell = ws.cell(row=5, column=2)
        # No fill means fgColor is either 0 or "00000000" (default)
        assert header_cell.fill.patternType is None or header_cell.fill.start_color.rgb in (
            "00000000", "0",
        )

    def test_no_fill_on_data_cells(self, config, asian_b_data):
        """Data cells have no fill and no borders."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        data_cell = ws.cell(row=6, column=2)
        assert data_cell.fill.patternType is None or data_cell.fill.start_color.rgb in (
            "00000000", "0",
        )

    def test_chart_title_below_data(self, config, asian_b_data):
        """Chart title is 2 rows below data row (row 8)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        title = str(ws.cell(row=8, column=1).value)
        assert "Asian" in title
        assert "White" in title or "2017-2023" in title

    def test_single_race_chart_created(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        assert len(ws._charts) == 1

    def test_three_races_three_charts(self, config, data_list):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, data_list, config)
        assert len(ws._charts) == 3

    def test_chart_has_one_series(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        chart = ws._charts[0]
        assert len(chart.series) == 1

    def test_chart_type_is_col(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        chart = ws._charts[0]
        assert chart.type == "col"
        assert chart.grouping == "clustered"

    def test_chart_gap_and_overlap(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        chart = ws._charts[0]
        assert chart.gapWidth == 219
        assert chart.overlap == -27

    def test_data_labels_enabled(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        chart = ws._charts[0]
        series = chart.series[0]
        assert series.dLbls is not None
        assert series.dLbls.showVal is True

    def test_no_descriptive_text(self, config, asian_b_data):
        """New format has no descriptive text."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        found_descriptive = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "age-adjusted" in str(row[0]).lower():
                found_descriptive = True
                break
        assert not found_descriptive, "Descriptive text should not be present in new format"

    def test_no_footnote(self, config, asian_b_data):
        """New format has no footnotes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        found_footnote = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "DATA SOURCE" in str(row[0]):
                found_footnote = True
                break
        assert not found_footnote, "Footnote should not be present in new format"

    def test_chart_has_no_legend(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        chart = ws._charts[0]
        assert chart.legend is None

    def test_y_axis_has_title(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        chart = ws._charts[0]
        assert chart.y_axis.title is not None
        assert "per 100,000" in str(chart.y_axis.title)

    def test_series_fill_is_set(self, config, asian_b_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_chart_set_b_sheet(ws, [asian_b_data], config)
        chart = ws._charts[0]
        assert chart.series[0].graphicalProperties.solidFill is not None


# ---------------------------------------------------------------------------
# WorkbookBuilder.add_chart_set_b integration tests
# ---------------------------------------------------------------------------

class TestAddChartSetBIntegration:
    """Test the add_chart_set_b method via WorkbookBuilder."""

    def test_empty_data_list_no_sheet(self, builder):
        builder.add_chart_set_b([])
        assert "OUTPUT-2" not in builder.wb.sheetnames

    def test_creates_output_2_sheet(self, builder, data_list):
        builder.add_chart_set_b(data_list)
        assert "OUTPUT-2" in builder.wb.sheetnames

    def test_sheet_has_three_charts(self, builder, data_list):
        builder.add_chart_set_b(data_list)
        ws = builder.wb["OUTPUT-2"]
        assert len(ws._charts) == 3

    def test_save_bytes_produces_valid_xlsx(self, builder, data_list):
        builder.add_chart_set_b(data_list)
        data = builder.save_bytes()
        assert data[:2] == b"PK"
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        assert "OUTPUT-2" in loaded.sheetnames

    def test_save_and_reload_preserves_data(self, builder, asian_b_data):
        builder.add_chart_set_b([asian_b_data])
        data = builder.save_bytes()
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        ws = loaded["OUTPUT-2"]
        # Data at row 6 cols B/C/D
        assert ws.cell(row=6, column=2).value == 110.5
        assert ws.cell(row=6, column=3).value == 125.3
        assert ws.cell(row=6, column=4).value == 128.8

    def test_all_three_race_blocks_have_data(self, builder, data_list):
        builder.add_chart_set_b(data_list)
        ws = builder.wb["OUTPUT-2"]
        all_values = []
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=4, values_only=True):
            all_values.extend([str(v) for v in row if v is not None])
        joined = " ".join(all_values)
        assert "Asian" in joined
        assert "Black" in joined
        assert "Latinx" in joined
