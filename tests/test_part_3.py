"""Tests for Part 3: Sex- and Race-Stratified."""

from __future__ import annotations

import io

import openpyxl
import pytest

from autochart.builder.workbook import WorkbookBuilder
from autochart.charts.part_3 import build_part_3_sheet
from autochart.config import ChartConfig, Part3Data, RateComparison


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def config() -> ChartConfig:
    return ChartConfig(
        disease_name="Cancer Mortality",
        rate_unit="per 100,000 residents",
        rate_denominator=100_000,
        data_source="DATA SOURCE: Boston resident deaths, Massachusetts Department of Public Health",
        years="2017-2023",
    )


def _make_comparison(
    group_name: str,
    group_rate: float,
    reference_name: str,
    reference_rate: float,
    p_value: float | None = None,
) -> RateComparison:
    return RateComparison(
        group_name=group_name,
        group_rate=group_rate,
        reference_name=reference_name,
        reference_rate=reference_rate,
        p_value=p_value,
    )


@pytest.fixture()
def part_3_data() -> Part3Data:
    return Part3Data(
        female_comparisons=[
            _make_comparison("Asian", 87.9, "White", 108.2, p_value=0.01),
            _make_comparison("Black", 127.7, "White", 108.2, p_value=0.005),
            _make_comparison("Latinx", 88.5, "White", 108.2, p_value=0.005),
        ],
        male_comparisons=[
            _make_comparison("Asian", 141.1, "White", 148.0, p_value=0.35),
            _make_comparison("Black", 188.5, "White", 148.0, p_value=0.0001),
            _make_comparison("Latinx", 113.9, "White", 148.0, p_value=0.0001),
        ],
        female_boston_rate=111.5,
        male_boston_rate=150.8,
    )


@pytest.fixture()
def builder(config: ChartConfig) -> WorkbookBuilder:
    return WorkbookBuilder(config)


# ---------------------------------------------------------------------------
# build_part_3_sheet -- direct function tests
# ---------------------------------------------------------------------------

class TestBuildPart3Sheet:
    """Test the build_part_3_sheet function directly."""

    def test_section_header(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        assert ws.cell(row=1, column=1).value == "All Cancer Mortality"

    def test_chart_title(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        title = str(ws.cell(row=3, column=1).value)
        assert "by Sex and Race" in title
        assert "2017-2023" in title

    def test_data_table_headers(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        # Headers in row 4
        assert ws.cell(row=4, column=1).value == ""
        assert ws.cell(row=4, column=2).value == "Asian"
        assert ws.cell(row=4, column=3).value == "Black"
        assert ws.cell(row=4, column=4).value == "Latinx"
        assert ws.cell(row=4, column=5).value == "White"
        assert ws.cell(row=4, column=6).value == "Boston"

    def test_data_table_female_row(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        # Female row at row 5
        assert ws.cell(row=5, column=1).value == "Female"
        assert ws.cell(row=5, column=2).value == 87.9   # Asian female
        assert ws.cell(row=5, column=3).value == 127.7  # Black female
        assert ws.cell(row=5, column=4).value == 88.5   # Latinx female
        assert ws.cell(row=5, column=5).value == 108.2  # White female
        assert ws.cell(row=5, column=6).value == 111.5  # Boston female

    def test_data_table_male_row(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        # Male row at row 6
        assert ws.cell(row=6, column=1).value == "Male"
        assert ws.cell(row=6, column=2).value == 141.1  # Asian male
        assert ws.cell(row=6, column=3).value == 188.5  # Black male
        assert ws.cell(row=6, column=4).value == 113.9  # Latinx male
        assert ws.cell(row=6, column=5).value == 148.0  # White male
        assert ws.cell(row=6, column=6).value == 150.8  # Boston male

    def test_chart_created(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        assert len(ws._charts) == 1

    def test_chart_has_two_series(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        chart = ws._charts[0]
        assert len(chart.series) == 2

    def test_chart_type_is_col(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        chart = ws._charts[0]
        assert chart.type == "col"
        assert chart.grouping == "clustered"

    def test_chart_gap_and_overlap(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        chart = ws._charts[0]
        assert chart.gapWidth == 219
        assert chart.overlap == -27

    def test_data_labels_enabled(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        chart = ws._charts[0]
        for series in chart.series:
            assert series.dLbls is not None
            assert series.dLbls.showVal is True

    def test_descriptive_text_present(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        found_descriptive = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "age-adjusted" in str(row[0]).lower():
                found_descriptive = True
                break
        assert found_descriptive, "Descriptive text not found in the sheet"

    def test_footnote_present(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        found_footnote = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1, values_only=True):
            if row[0] and "DATA SOURCE" in str(row[0]):
                found_footnote = True
                break
        assert found_footnote, "Footnote text not found in the sheet"

    def test_chart_has_no_legend(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        chart = ws._charts[0]
        assert chart.legend is None

    def test_y_axis_has_title(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        chart = ws._charts[0]
        assert chart.y_axis.title is not None
        assert "per 100,000" in str(chart.y_axis.title)

    def test_header_cells_have_gray_fill(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        header_cell = ws.cell(row=4, column=2)
        assert header_cell.fill.start_color.rgb == "00D9D9D9"

    def test_both_series_are_navy(self, config, part_3_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        build_part_3_sheet(ws, part_3_data, config)
        chart = ws._charts[0]
        for i, series in enumerate(chart.series):
            assert series.graphicalProperties.solidFill is not None, (
                f"Series {i} should have a navy solid fill"
            )


# ---------------------------------------------------------------------------
# WorkbookBuilder.add_part_3 integration tests
# ---------------------------------------------------------------------------

class TestAddPart3Integration:
    """Test the add_part_3 method via WorkbookBuilder."""

    def test_creates_output_4_sheet(self, builder, part_3_data):
        builder.add_part_3(part_3_data)
        assert "OUTPUT-4" in builder.wb.sheetnames

    def test_sheet_has_one_chart(self, builder, part_3_data):
        builder.add_part_3(part_3_data)
        ws = builder.wb["OUTPUT-4"]
        assert len(ws._charts) == 1

    def test_save_bytes_produces_valid_xlsx(self, builder, part_3_data):
        builder.add_part_3(part_3_data)
        data = builder.save_bytes()
        assert data[:2] == b"PK"
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        assert "OUTPUT-4" in loaded.sheetnames

    def test_save_and_reload_preserves_data(self, builder, part_3_data):
        builder.add_part_3(part_3_data)
        data = builder.save_bytes()
        loaded = openpyxl.load_workbook(io.BytesIO(data))
        ws = loaded["OUTPUT-4"]
        # Check female row values
        assert ws.cell(row=5, column=1).value == "Female"
        assert ws.cell(row=5, column=2).value == 87.9
        # Check male row values
        assert ws.cell(row=6, column=1).value == "Male"
        assert ws.cell(row=6, column=2).value == 141.1

    def test_multiple_calls_create_unique_sheets(self, builder, part_3_data):
        builder.add_part_3(part_3_data)
        builder.add_part_3(part_3_data)
        assert "OUTPUT-4" in builder.wb.sheetnames
        assert "OUTPUT-4_2" in builder.wb.sheetnames
