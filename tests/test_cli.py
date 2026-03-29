"""Tests for autochart.cli -- CLI entry point."""

from __future__ import annotations

import io
import os
import tempfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from autochart.cli import (
    _compute_chart_patches,
    _parse_chart_types,
    _run_generate,
    build_parser,
    main,
)
from autochart.config import (
    ChartConfig,
    ChartSetAData,
    ChartSetBData,
    ChartSetCData,
    ChartSetType,
    Part3Data,
    RateComparison,
)
from autochart.builder.postprocess import ChartPatch


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

EXAMPLES_PATH = str(
    Path(__file__).resolve().parent.parent / "examples" / "examples.xlsx"
)


@pytest.fixture()
def config() -> ChartConfig:
    return ChartConfig(
        disease_name="Cancer Mortality",
        rate_unit="per 100,000 residents",
        rate_denominator=100_000,
        data_source="DATA SOURCE: Test",
        years="2017-2023",
        demographics=["Asian", "Black", "Latinx", "White"],
        reference_group="White",
        geography="Boston",
    )


def _make_rate_comparison(
    group: str = "Asian",
    group_rate: float = 110.5,
    reference: str = "White",
    reference_rate: float = 130.6,
    p_value: float | None = 0.01,
) -> RateComparison:
    return RateComparison(
        group_name=group,
        group_rate=group_rate,
        reference_name=reference,
        reference_rate=reference_rate,
        p_value=p_value,
    )


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

class TestBuildParser:
    def test_parser_created(self):
        parser = build_parser()
        assert parser is not None

    def test_generate_subcommand_minimal(self):
        parser = build_parser()
        args = parser.parse_args([
            "generate", "input.xlsx",
            "--disease", "Cancer Mortality",
            "--years", "2017-2023",
        ])
        assert args.command == "generate"
        assert args.input_file == "input.xlsx"
        assert args.disease == "Cancer Mortality"
        assert args.years == "2017-2023"

    def test_defaults(self):
        parser = build_parser()
        args = parser.parse_args([
            "generate", "input.xlsx",
            "--disease", "Test",
            "--years", "2020-2024",
        ])
        assert args.output == "output.xlsx"
        assert args.rate_unit is None
        assert args.rate_denominator is None
        assert args.data_source is None
        assert args.charts == "all"
        assert args.geography == "Boston"
        assert args.reference_group == "White"
        assert args.demographics == "Asian,Black,Latinx,White"
        assert args.no_auto is False

    def test_all_arguments(self):
        parser = build_parser()
        args = parser.parse_args([
            "generate", "input.xlsx",
            "-o", "custom_output.xlsx",
            "--disease", "Heart Disease",
            "--rate-unit", "per 10,000 residents",
            "--rate-denominator", "10000",
            "--data-source", "DATA SOURCE: CDC",
            "--years", "2018-2024",
            "--charts", "a,b",
            "--geography", "New York",
            "--reference-group", "White",
            "--demographics", "Asian,Black,Hispanic,White",
        ])
        assert args.output == "custom_output.xlsx"
        assert args.disease == "Heart Disease"
        assert args.rate_unit == "per 10,000 residents"
        assert args.rate_denominator == 10000
        assert args.data_source == "DATA SOURCE: CDC"
        assert args.years == "2018-2024"
        assert args.charts == "a,b"
        assert args.geography == "New York"
        assert args.reference_group == "White"
        assert args.demographics == "Asian,Black,Hispanic,White"

    def test_no_subcommand_exits(self):
        """When no subcommand is given, main() should print help and exit."""
        with pytest.raises(SystemExit) as exc_info:
            main([])
        assert exc_info.value.code == 1

    def test_disease_not_required(self):
        """--disease is optional (default=None) for auto-detection."""
        parser = build_parser()
        args = parser.parse_args(["generate", "input.xlsx"])
        assert args.disease is None

    def test_years_not_required(self):
        """--years is optional (default=None) for auto-detection."""
        parser = build_parser()
        args = parser.parse_args(["generate", "input.xlsx"])
        assert args.years is None

    def test_no_auto_flag(self):
        """--no-auto flag is parsed correctly."""
        parser = build_parser()
        args = parser.parse_args(["generate", "input.xlsx", "--no-auto"])
        assert args.no_auto is True

    def test_no_auto_flag_default(self):
        """--no-auto defaults to False."""
        parser = build_parser()
        args = parser.parse_args(["generate", "input.xlsx"])
        assert args.no_auto is False


# ---------------------------------------------------------------------------
# Chart type parsing
# ---------------------------------------------------------------------------

class TestParseChartTypes:
    def test_all(self):
        result = _parse_chart_types("all")
        assert result == [
            ChartSetType.A, ChartSetType.B, ChartSetType.C, ChartSetType.PART_3
        ]

    def test_single_a(self):
        assert _parse_chart_types("a") == [ChartSetType.A]

    def test_single_b(self):
        assert _parse_chart_types("b") == [ChartSetType.B]

    def test_single_c(self):
        assert _parse_chart_types("c") == [ChartSetType.C]

    def test_single_part3(self):
        assert _parse_chart_types("part3") == [ChartSetType.PART_3]

    def test_multiple(self):
        result = _parse_chart_types("a,b,c")
        assert result == [ChartSetType.A, ChartSetType.B, ChartSetType.C]

    def test_multiple_with_spaces(self):
        result = _parse_chart_types("a, b, part3")
        assert result == [ChartSetType.A, ChartSetType.B, ChartSetType.PART_3]

    def test_deduplication(self):
        result = _parse_chart_types("a,a,b")
        assert result == [ChartSetType.A, ChartSetType.B]

    def test_unknown_raises(self):
        with pytest.raises(ValueError, match="Unknown chart type"):
            _parse_chart_types("x")

    def test_part_3_alias(self):
        assert _parse_chart_types("part_3") == [ChartSetType.PART_3]

    def test_case_insensitive(self):
        result = _parse_chart_types("A,B,C")
        assert result == [ChartSetType.A, ChartSetType.B, ChartSetType.C]


# ---------------------------------------------------------------------------
# Chart patch computation
# ---------------------------------------------------------------------------

class TestComputeChartPatches:
    def test_chart_set_b_pattern_fill_on_white(self, config):
        """Chart Set B: White bar (index 1) should always get pattern fill."""
        comp = _make_rate_comparison(p_value=0.5)  # not significant
        b_data = [ChartSetBData(race_name="Asian", comparison=comp, boston_overall_rate=128.8)]

        by_type = {ChartSetType.B: b_data}
        patches = _compute_chart_patches(by_type, [ChartSetType.B], config)

        assert len(patches) == 1
        assert patches[0].chart_index == 1
        assert 1 in patches[0].pattern_fill_points
        assert patches[0].asterisk_points == []

    def test_chart_set_b_significant_gets_asterisk(self, config):
        """Chart Set B: significant comparison -> asterisk on race bar."""
        comp = _make_rate_comparison(p_value=0.001)  # significant
        b_data = [ChartSetBData(race_name="Black", comparison=comp, boston_overall_rate=128.8)]

        by_type = {ChartSetType.B: b_data}
        patches = _compute_chart_patches(by_type, [ChartSetType.B], config)

        assert len(patches) == 1
        assert 0 in patches[0].asterisk_points

    def test_chart_set_b_multiple_races(self, config):
        """Chart Set B with 3 races produces 3 patches."""
        races = ["Asian", "Black", "Latinx"]
        b_data = [
            ChartSetBData(
                race_name=r,
                comparison=_make_rate_comparison(group=r, p_value=0.01),
                boston_overall_rate=128.8,
            )
            for r in races
        ]

        by_type = {ChartSetType.B: b_data}
        patches = _compute_chart_patches(by_type, [ChartSetType.B], config)

        assert len(patches) == 3
        for i, p in enumerate(patches):
            assert p.chart_index == i + 1

    def test_chart_set_c_white_pattern_fill(self, config):
        """Chart Set C: White bar gets pattern fill at index = num_races."""
        comparisons = [
            _make_rate_comparison(group="Asian", p_value=0.5),
            _make_rate_comparison(group="Black", p_value=0.5),
            _make_rate_comparison(group="Latinx", p_value=0.5),
        ]
        c_data = ChartSetCData(comparisons=comparisons, boston_overall_rate=128.8)

        by_type = {ChartSetType.C: [c_data]}
        patches = _compute_chart_patches(by_type, [ChartSetType.C], config)

        assert len(patches) == 1
        assert patches[0].pattern_fill_points == [3]  # 3 races -> White at index 3

    def test_chart_set_c_significant_asterisks(self, config):
        """Chart Set C: significant race bars get asterisks."""
        comparisons = [
            _make_rate_comparison(group="Asian", p_value=0.001),
            _make_rate_comparison(group="Black", p_value=0.5),
            _make_rate_comparison(group="Latinx", p_value=0.02),
        ]
        c_data = ChartSetCData(comparisons=comparisons, boston_overall_rate=128.8)

        by_type = {ChartSetType.C: [c_data]}
        patches = _compute_chart_patches(by_type, [ChartSetType.C], config)

        assert 0 in patches[0].asterisk_points  # Asian significant
        assert 1 not in patches[0].asterisk_points  # Black not significant
        assert 2 in patches[0].asterisk_points  # Latinx significant

    def test_part_3_two_series_patches(self, config):
        """Part 3 produces 2 patches (female + male series)."""
        female = [
            _make_rate_comparison(group="Asian", p_value=0.01),
            _make_rate_comparison(group="Black", p_value=0.5),
            _make_rate_comparison(group="Latinx", p_value=0.02),
        ]
        male = [
            _make_rate_comparison(group="Asian", p_value=0.5),
            _make_rate_comparison(group="Black", p_value=0.001),
            _make_rate_comparison(group="Latinx", p_value=0.5),
        ]
        p3_data = Part3Data(
            female_comparisons=female,
            male_comparisons=male,
            female_boston_rate=111.0,
            male_boston_rate=155.0,
        )

        by_type = {ChartSetType.PART_3: [p3_data]}
        patches = _compute_chart_patches(by_type, [ChartSetType.PART_3], config)

        assert len(patches) == 2
        # Both patches target the same chart
        assert patches[0].chart_index == patches[1].chart_index == 1
        # Female series (0) and male series (1)
        assert patches[0].series_index == 0
        assert patches[1].series_index == 1
        # White bar pattern fills
        assert patches[0].pattern_fill_points == [3]
        assert patches[1].pattern_fill_points == [3]

    def test_chart_set_a_significant_asterisks(self, config):
        """Chart Set A: significant comparisons get asterisks on race series."""
        a_data = ChartSetAData(
            race_name="Asian",
            boston=_make_rate_comparison(p_value=0.01),
            female=_make_rate_comparison(p_value=0.5),
            male=_make_rate_comparison(p_value=0.02),
            boston_overall_rate=128.8,
            female_overall_rate=111.0,
            male_overall_rate=155.0,
        )

        by_type = {ChartSetType.A: [a_data]}
        patches = _compute_chart_patches(by_type, [ChartSetType.A], config)

        assert len(patches) == 1
        assert patches[0].chart_index == 1
        assert 0 in patches[0].asterisk_points  # boston significant
        assert 1 not in patches[0].asterisk_points  # female not significant
        assert 2 in patches[0].asterisk_points  # male significant

    def test_chart_set_a_no_significance_no_patch(self, config):
        """Chart Set A with no significant comparisons creates no patches."""
        a_data = ChartSetAData(
            race_name="Asian",
            boston=_make_rate_comparison(p_value=0.5),
            female=_make_rate_comparison(p_value=0.5),
            male=_make_rate_comparison(p_value=0.5),
            boston_overall_rate=128.8,
            female_overall_rate=111.0,
            male_overall_rate=155.0,
        )

        by_type = {ChartSetType.A: [a_data]}
        patches = _compute_chart_patches(by_type, [ChartSetType.A], config)

        assert len(patches) == 0

    def test_chart_numbering_across_types(self, config):
        """Chart indices increment across multiple chart set types."""
        # Set A: 3 races = 3 charts (all significant to produce patches)
        a_data = [
            ChartSetAData(
                race_name=r,
                boston=_make_rate_comparison(group=r, p_value=0.01),
                female=_make_rate_comparison(group=r, p_value=0.5),
                male=_make_rate_comparison(group=r, p_value=0.5),
                boston_overall_rate=128.8,
                female_overall_rate=111.0,
                male_overall_rate=155.0,
            )
            for r in ["Asian", "Black", "Latinx"]
        ]

        # Set B: 3 races = 3 charts
        b_data = [
            ChartSetBData(
                race_name=r,
                comparison=_make_rate_comparison(group=r, p_value=0.01),
                boston_overall_rate=128.8,
            )
            for r in ["Asian", "Black", "Latinx"]
        ]

        by_type = {ChartSetType.A: a_data, ChartSetType.B: b_data}
        patches = _compute_chart_patches(
            by_type, [ChartSetType.A, ChartSetType.B], config
        )

        # A produces 3 patches (charts 1, 2, 3), B produces 3 patches (charts 4, 5, 6)
        a_patches = [p for p in patches if p.pattern_fill_points == []]
        b_patches = [p for p in patches if 1 in p.pattern_fill_points]

        assert len(a_patches) == 3
        assert len(b_patches) == 3
        assert sorted(p.chart_index for p in a_patches) == [1, 2, 3]
        assert sorted(p.chart_index for p in b_patches) == [4, 5, 6]

    def test_missing_chart_type_skipped(self, config):
        """Requesting a type not in by_type produces no patches for it."""
        by_type: dict[ChartSetType, list] = {}
        patches = _compute_chart_patches(by_type, [ChartSetType.A], config)
        assert patches == []


# ---------------------------------------------------------------------------
# End-to-end: main() with examples file
# ---------------------------------------------------------------------------

class TestEndToEnd:
    @pytest.fixture()
    def output_path(self, tmp_path):
        return str(tmp_path / "test_output.xlsx")

    @pytest.mark.skipif(
        not Path(EXAMPLES_PATH).exists(),
        reason="examples.xlsx not found",
    )
    def test_generate_all_charts(self, output_path, capsys):
        """End-to-end: generate all charts from examples.xlsx."""
        main([
            "generate", EXAMPLES_PATH,
            "-o", output_path,
            "--disease", "Cancer Mortality",
            "--years", "2018-2024",
            "--data-source", "DATA SOURCE: Test",
        ])

        assert Path(output_path).exists()
        assert Path(output_path).stat().st_size > 0

        # Verify it's a valid xlsx
        wb = openpyxl.load_workbook(output_path)
        assert len(wb.sheetnames) > 0
        wb.close()

        # Check stdout
        captured = capsys.readouterr()
        assert "Generation complete" in captured.out

    @pytest.mark.skipif(
        not Path(EXAMPLES_PATH).exists(),
        reason="examples.xlsx not found",
    )
    def test_generate_specific_chart_types(self, output_path, capsys):
        """End-to-end: generate only Chart Set A."""
        main([
            "generate", EXAMPLES_PATH,
            "-o", output_path,
            "--disease", "Cancer Mortality",
            "--years", "2018-2024",
            "--charts", "a",
        ])

        assert Path(output_path).exists()
        captured = capsys.readouterr()
        assert "Race vs Rest of City" in captured.out

    def test_missing_input_file_exits(self, capsys):
        """Providing a non-existent input file should exit with error."""
        with pytest.raises(SystemExit) as exc_info:
            main([
                "generate", "/nonexistent/file.xlsx",
                "--disease", "Test",
                "--years", "2020-2024",
            ])
        assert exc_info.value.code == 1

    def test_non_xlsx_file_exits(self, tmp_path, capsys):
        """Providing a non-.xlsx file should exit with error."""
        csv_file = tmp_path / "data.csv"
        csv_file.write_text("a,b,c")

        with pytest.raises(SystemExit) as exc_info:
            main([
                "generate", str(csv_file),
                "--disease", "Test",
                "--years", "2020-2024",
            ])
        assert exc_info.value.code == 1

    def test_invalid_chart_type_exits(self, tmp_path, capsys):
        """Invalid chart type in --charts should exit with error."""
        # Create a dummy xlsx
        wb = openpyxl.Workbook()
        dummy_path = str(tmp_path / "dummy.xlsx")
        wb.save(dummy_path)
        wb.close()

        with pytest.raises(SystemExit) as exc_info:
            main([
                "generate", dummy_path,
                "--disease", "Test",
                "--years", "2020-2024",
                "--charts", "invalid_type",
            ])
        assert exc_info.value.code == 1

    @pytest.mark.skipif(
        not Path(EXAMPLES_PATH).exists(),
        reason="examples.xlsx not found",
    )
    def test_auto_extract_end_to_end(self, output_path, capsys):
        """Zero-config: generate without --disease/--years using auto-detection."""
        main([
            "generate", EXAMPLES_PATH,
            "-o", output_path,
        ])

        assert Path(output_path).exists()
        assert Path(output_path).stat().st_size > 0

        # Verify it's a valid xlsx
        wb = openpyxl.load_workbook(output_path)
        assert len(wb.sheetnames) > 0
        wb.close()

        # Check stdout shows auto-detected config
        captured = capsys.readouterr()
        assert "Auto-detected configuration" in captured.out
        assert "Generation complete" in captured.out

    @pytest.mark.skipif(
        not Path(EXAMPLES_PATH).exists(),
        reason="examples.xlsx not found",
    )
    def test_override_beats_extraction(self, output_path, capsys):
        """CLI --disease override should appear in output, beating auto-detection."""
        main([
            "generate", EXAMPLES_PATH,
            "-o", output_path,
            "--disease", "Custom Disease Name",
        ])

        captured = capsys.readouterr()
        assert "Custom Disease Name" in captured.out

    @pytest.mark.skipif(
        not Path(EXAMPLES_PATH).exists(),
        reason="examples.xlsx not found",
    )
    def test_no_auto_requires_disease(self, output_path, capsys):
        """--no-auto without --disease should error."""
        with pytest.raises(SystemExit) as exc_info:
            main([
                "generate", EXAMPLES_PATH,
                "-o", output_path,
                "--no-auto",
                "--years", "2018-2024",
            ])
        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "--disease is required" in captured.err

    @pytest.mark.skipif(
        not Path(EXAMPLES_PATH).exists(),
        reason="examples.xlsx not found",
    )
    def test_backward_compat(self, output_path, capsys):
        """Old CLI with --disease and --years still works (backward compat)."""
        main([
            "generate", EXAMPLES_PATH,
            "-o", output_path,
            "--disease", "Cancer Mortality",
            "--years", "2018-2024",
            "--data-source", "DATA SOURCE: Test",
        ])

        assert Path(output_path).exists()
        captured = capsys.readouterr()
        assert "Generation complete" in captured.out
        assert "Cancer Mortality" in captured.out
