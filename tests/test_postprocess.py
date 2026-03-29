"""Tests for autochart.builder.postprocess -- OOXML post-processing pipeline."""

from __future__ import annotations

import io
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference

from autochart.builder.postprocess import (
    ChartPatch,
    postprocess_xlsx,
    _apply_montserrat_font,
    _apply_pattern_fill_to_point,
    _apply_asterisk_to_point,
)
from autochart.charts.ooxml import NSMAP, _qn


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx_with_chart(num_points: int = 5, num_charts: int = 1) -> bytes:
    """Build a minimal .xlsx with an openpyxl bar chart for testing.

    Returns the raw .xlsx bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write some data for the chart
    ws.append(["Category", "Value"])
    for i in range(num_points):
        ws.append([f"Cat {i}", (i + 1) * 10])

    for _ in range(num_charts):
        chart = BarChart()
        chart.type = "col"
        chart.title = "Test Chart"
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=num_points + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=num_points + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _extract_chart_xml(xlsx_bytes: bytes, chart_index: int = 1) -> bytes:
    """Extract a chart XML file from an .xlsx archive."""
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zf:
        return zf.read(f"xl/charts/chart{chart_index}.xml")


def _minimal_chart_xml(
    num_points: int = 5, with_dlbls: bool = False
) -> bytes:
    """Build minimal chart XML bytes for unit-testing patch functions."""
    for prefix, uri in NSMAP.items():
        ET.register_namespace(prefix, uri)

    root = ET.Element(_qn("c", "chartSpace"))
    chart = ET.SubElement(root, _qn("c", "chart"))
    plot = ET.SubElement(chart, _qn("c", "plotArea"))
    bar = ET.SubElement(plot, _qn("c", "barChart"))

    ser = ET.SubElement(bar, _qn("c", "ser"))
    ET.SubElement(ser, _qn("c", "idx"), attrib={"val": "0"})
    ET.SubElement(ser, _qn("c", "order"), attrib={"val": "0"})

    if with_dlbls:
        ET.SubElement(ser, _qn("c", "dLbls"))

    # Category axis
    cat_ax = ET.SubElement(plot, _qn("c", "catAx"))
    ET.SubElement(cat_ax, _qn("c", "axId"), attrib={"val": "1"})

    # Value axis
    val_ax = ET.SubElement(plot, _qn("c", "valAx"))
    ET.SubElement(val_ax, _qn("c", "axId"), attrib={"val": "2"})

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


# ---------------------------------------------------------------------------
# Tests: Montserrat font
# ---------------------------------------------------------------------------

class TestMontserratFont:
    """Test that _apply_montserrat_font sets Montserrat on chart elements."""

    def test_sets_font_on_data_labels(self):
        """Data labels should get Montserrat latin and cs fonts."""
        chart_xml = _minimal_chart_xml(with_dlbls=True)
        result = _apply_montserrat_font(chart_xml)
        root = ET.fromstring(result)

        dlbls = root.find(f".//{_qn('c', 'dLbls')}")
        assert dlbls is not None

        txPr = dlbls.find(_qn("c", "txPr"))
        assert txPr is not None

        # Find default run properties
        defRPr = txPr.find(
            f"{_qn('a', 'p')}/{_qn('a', 'pPr')}/{_qn('a', 'defRPr')}"
        )
        assert defRPr is not None

        latin = defRPr.find(_qn("a", "latin"))
        assert latin is not None
        assert latin.get("typeface") == "Montserrat"

        cs = defRPr.find(_qn("a", "cs"))
        assert cs is not None
        assert cs.get("typeface") == "Montserrat"

    def test_data_label_font_size_900(self):
        """Data labels should use sz=900 (9 pt)."""
        chart_xml = _minimal_chart_xml(with_dlbls=True)
        result = _apply_montserrat_font(chart_xml)
        root = ET.fromstring(result)

        defRPr = root.find(
            f".//{_qn('c', 'dLbls')}/{_qn('c', 'txPr')}"
            f"/{_qn('a', 'p')}/{_qn('a', 'pPr')}/{_qn('a', 'defRPr')}"
        )
        assert defRPr.get("sz") == "900"

    def test_data_label_color_scheme(self):
        """Data labels should get tx1 colour with lumMod/lumOff."""
        chart_xml = _minimal_chart_xml(with_dlbls=True)
        result = _apply_montserrat_font(chart_xml)
        root = ET.fromstring(result)

        defRPr = root.find(
            f".//{_qn('c', 'dLbls')}/{_qn('c', 'txPr')}"
            f"/{_qn('a', 'p')}/{_qn('a', 'pPr')}/{_qn('a', 'defRPr')}"
        )
        solidFill = defRPr.find(_qn("a", "solidFill"))
        assert solidFill is not None

        schemeClr = solidFill.find(_qn("a", "schemeClr"))
        assert schemeClr is not None
        assert schemeClr.get("val") == "tx1"

        lumMod = schemeClr.find(_qn("a", "lumMod"))
        assert lumMod is not None
        assert lumMod.get("val") == "75000"

        lumOff = schemeClr.find(_qn("a", "lumOff"))
        assert lumOff is not None
        assert lumOff.get("val") == "25000"

    def test_sets_font_on_cat_axis(self):
        """Category axis tick labels should get Montserrat."""
        chart_xml = _minimal_chart_xml()
        result = _apply_montserrat_font(chart_xml)
        root = ET.fromstring(result)

        catAx = root.find(f".//{_qn('c', 'catAx')}")
        txPr = catAx.find(_qn("c", "txPr"))
        assert txPr is not None

        defRPr = txPr.find(
            f"{_qn('a', 'p')}/{_qn('a', 'pPr')}/{_qn('a', 'defRPr')}"
        )
        latin = defRPr.find(_qn("a", "latin"))
        assert latin is not None
        assert latin.get("typeface") == "Montserrat"

    def test_sets_font_on_val_axis(self):
        """Value axis tick labels should get Montserrat."""
        chart_xml = _minimal_chart_xml()
        result = _apply_montserrat_font(chart_xml)
        root = ET.fromstring(result)

        valAx = root.find(f".//{_qn('c', 'valAx')}")
        txPr = valAx.find(_qn("c", "txPr"))
        assert txPr is not None

        defRPr = txPr.find(
            f"{_qn('a', 'p')}/{_qn('a', 'pPr')}/{_qn('a', 'defRPr')}"
        )
        latin = defRPr.find(_qn("a", "latin"))
        assert latin is not None
        assert latin.get("typeface") == "Montserrat"

    def test_sets_font_on_title(self):
        """Chart/axis title run properties should get Montserrat."""
        # Build chart XML with a title containing a run
        for prefix, uri in NSMAP.items():
            ET.register_namespace(prefix, uri)

        root = ET.Element(_qn("c", "chartSpace"))
        chart = ET.SubElement(root, _qn("c", "chart"))
        title = ET.SubElement(chart, _qn("c", "title"))
        tx = ET.SubElement(title, _qn("c", "tx"))
        rich = ET.SubElement(tx, _qn("c", "rich"))
        ET.SubElement(rich, _qn("a", "bodyPr"))
        ET.SubElement(rich, _qn("a", "lstStyle"))
        p = ET.SubElement(rich, _qn("a", "p"))
        r = ET.SubElement(p, _qn("a", "r"))
        rPr = ET.SubElement(r, _qn("a", "rPr"), attrib={"lang": "en-US"})
        t = ET.SubElement(r, _qn("a", "t"))
        t.text = "My Title"

        chart_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        result = _apply_montserrat_font(chart_bytes)
        patched = ET.fromstring(result)

        rPr_patched = patched.find(f".//{_qn('a', 'rPr')}")
        latin = rPr_patched.find(_qn("a", "latin"))
        assert latin is not None
        assert latin.get("typeface") == "Montserrat"

    def test_idempotent(self):
        """Applying Montserrat twice should produce the same result."""
        chart_xml = _minimal_chart_xml(with_dlbls=True)
        result1 = _apply_montserrat_font(chart_xml)
        result2 = _apply_montserrat_font(result1)

        root1 = ET.fromstring(result1)
        root2 = ET.fromstring(result2)

        # Both should have exactly one latin element per defRPr
        for defRPr in root2.iter(_qn("a", "defRPr")):
            latins = defRPr.findall(_qn("a", "latin"))
            assert len(latins) == 1, "Duplicate latin elements after re-apply"


# ---------------------------------------------------------------------------
# Tests: Pattern fill
# ---------------------------------------------------------------------------

class TestPatternFill:
    """Test pattern fill injection on specific data points."""

    def test_injects_pattern_fill(self):
        """A dPt with pattFill should be created at the given index."""
        root = ET.fromstring(_minimal_chart_xml())
        _apply_pattern_fill_to_point(root, series_idx=0, point_idx=1)

        dpt = root.find(f".//{_qn('c', 'dPt')}")
        assert dpt is not None

        idx = dpt.find(_qn("c", "idx"))
        assert idx is not None
        assert idx.get("val") == "1"

        pattFill = dpt.find(f".//{_qn('a', 'pattFill')}")
        assert pattFill is not None
        assert pattFill.get("prst") == "wdDnDiag"

    def test_replaces_existing_solid_fill(self):
        """If a dPt already has a solidFill, it should be replaced."""
        root = ET.fromstring(_minimal_chart_xml())

        # Pre-create a dPt with solidFill
        ser = root.find(f".//{_qn('c', 'ser')}")
        dpt = ET.SubElement(ser, _qn("c", "dPt"))
        ET.SubElement(dpt, _qn("c", "idx"), attrib={"val": "2"})
        spPr = ET.SubElement(dpt, _qn("c", "spPr"))
        ET.SubElement(spPr, _qn("a", "solidFill"))

        _apply_pattern_fill_to_point(root, series_idx=0, point_idx=2)

        # solidFill should be gone
        assert spPr.find(_qn("a", "solidFill")) is None
        # pattFill should be present
        assert spPr.find(_qn("a", "pattFill")) is not None

    def test_no_op_for_missing_series(self):
        """Should silently do nothing if the series doesn't exist."""
        root = ET.fromstring(_minimal_chart_xml())
        # Series 99 doesn't exist
        _apply_pattern_fill_to_point(root, series_idx=99, point_idx=0)
        assert root.find(f".//{_qn('c', 'dPt')}") is None

    def test_multiple_points(self):
        """Multiple data points can receive pattern fills."""
        root = ET.fromstring(_minimal_chart_xml())
        _apply_pattern_fill_to_point(root, 0, 3)
        _apply_pattern_fill_to_point(root, 0, 8)

        dpts = list(root.iter(_qn("c", "dPt")))
        assert len(dpts) == 2

        indices = {
            dpt.find(_qn("c", "idx")).get("val") for dpt in dpts
        }
        assert indices == {"3", "8"}


# ---------------------------------------------------------------------------
# Tests: Asterisk data labels
# ---------------------------------------------------------------------------

class TestAsteriskDataLabels:
    """Test asterisk data label injection."""

    def test_injects_asterisk_label(self):
        """A dLbl with asterisk run should appear for the given point."""
        root = ET.fromstring(_minimal_chart_xml(with_dlbls=True))
        _apply_asterisk_to_point(root, series_idx=0, point_idx=2)

        dlbl = root.find(f".//{_qn('c', 'dLbl')}")
        assert dlbl is not None

        idx = dlbl.find(_qn("c", "idx"))
        assert idx.get("val") == "2"

        # Check for asterisk text run
        rich = dlbl.find(f"{_qn('c', 'tx')}/{_qn('c', 'rich')}")
        assert rich is not None

        para = rich.find(_qn("a", "p"))
        run = para.find(_qn("a", "r"))
        t = run.find(_qn("a", "t"))
        assert t.text == "*"

    def test_creates_dlbls_if_missing(self):
        """If no <c:dLbls> exists, one should be created."""
        root = ET.fromstring(_minimal_chart_xml(with_dlbls=False))
        _apply_asterisk_to_point(root, series_idx=0, point_idx=0)

        ser = root.find(f".//{_qn('c', 'ser')}")
        dlbls = ser.find(_qn("c", "dLbls"))
        assert dlbls is not None

        dlbl = dlbls.find(_qn("c", "dLbl"))
        assert dlbl is not None

    def test_replaces_existing_dlbl_for_same_point(self):
        """Re-applying asterisk to the same point should not duplicate."""
        root = ET.fromstring(_minimal_chart_xml(with_dlbls=True))
        _apply_asterisk_to_point(root, 0, 5)
        _apply_asterisk_to_point(root, 0, 5)

        dlbls = root.find(f".//{_qn('c', 'dLbls')}")
        dlbl_list = dlbls.findall(_qn("c", "dLbl"))
        # Should have exactly one dLbl for point 5
        matching = [
            d
            for d in dlbl_list
            if d.find(_qn("c", "idx")).get("val") == "5"
        ]
        assert len(matching) == 1

    def test_no_op_for_missing_series(self):
        """Should silently do nothing if the series doesn't exist."""
        root = ET.fromstring(_minimal_chart_xml())
        _apply_asterisk_to_point(root, series_idx=99, point_idx=0)
        assert root.find(f".//{_qn('c', 'dLbl')}") is None


# ---------------------------------------------------------------------------
# Tests: Full postprocess_xlsx pipeline
# ---------------------------------------------------------------------------

class TestPostprocessXlsx:
    """Integration tests for the full postprocess_xlsx pipeline."""

    def test_output_is_valid_xlsx(self):
        """The output bytes should be a valid .xlsx (ZIP) file."""
        xlsx_bytes = _make_xlsx_with_chart()
        result = postprocess_xlsx(xlsx_bytes, [])
        # Should open as a valid ZIP
        with zipfile.ZipFile(io.BytesIO(result), "r") as zf:
            names = zf.namelist()
            assert "[Content_Types].xml" in names

    def test_output_opens_with_openpyxl(self):
        """The output should be loadable by openpyxl."""
        xlsx_bytes = _make_xlsx_with_chart()
        result = postprocess_xlsx(xlsx_bytes, [])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) >= 1

    def test_montserrat_applied_to_chart(self):
        """After post-processing, chart XML should contain Montserrat."""
        xlsx_bytes = _make_xlsx_with_chart()
        result = postprocess_xlsx(xlsx_bytes, [])

        chart_xml = _extract_chart_xml(result, chart_index=1)
        root = ET.fromstring(chart_xml)

        # At least one element should reference Montserrat
        found = False
        for latin in root.iter(_qn("a", "latin")):
            if latin.get("typeface") == "Montserrat":
                found = True
                break
        assert found, "Montserrat font not found in post-processed chart XML"

    def test_pattern_fill_applied(self):
        """ChartPatch with pattern_fill_points should inject pattFill."""
        xlsx_bytes = _make_xlsx_with_chart(num_points=5)
        patches = [
            ChartPatch(chart_index=1, pattern_fill_points=[1, 3]),
        ]
        result = postprocess_xlsx(xlsx_bytes, patches)

        chart_xml = _extract_chart_xml(result, chart_index=1)
        root = ET.fromstring(chart_xml)

        dpts = list(root.iter(_qn("c", "dPt")))
        assert len(dpts) >= 2

        for dpt in dpts:
            patt = dpt.find(f".//{_qn('a', 'pattFill')}")
            assert patt is not None

    def test_asterisk_applied(self):
        """ChartPatch with asterisk_points should inject dLbl."""
        xlsx_bytes = _make_xlsx_with_chart(num_points=5)
        patches = [
            ChartPatch(chart_index=1, asterisk_points=[2]),
        ]
        result = postprocess_xlsx(xlsx_bytes, patches)

        chart_xml = _extract_chart_xml(result, chart_index=1)
        root = ET.fromstring(chart_xml)

        dlbl = root.find(f".//{_qn('c', 'dLbl')}")
        assert dlbl is not None
        idx = dlbl.find(_qn("c", "idx"))
        assert idx.get("val") == "2"

    def test_combined_patches(self):
        """Pattern fill and asterisk can be applied to the same chart."""
        xlsx_bytes = _make_xlsx_with_chart(num_points=5)
        patches = [
            ChartPatch(
                chart_index=1,
                pattern_fill_points=[1],
                asterisk_points=[0, 2],
            ),
        ]
        result = postprocess_xlsx(xlsx_bytes, patches)

        chart_xml = _extract_chart_xml(result, chart_index=1)
        root = ET.fromstring(chart_xml)

        # Pattern fill
        dpt = root.find(f".//{_qn('c', 'dPt')}")
        assert dpt is not None
        assert dpt.find(f".//{_qn('a', 'pattFill')}") is not None

        # Asterisk labels
        dlbls = list(root.iter(_qn("c", "dLbl")))
        assert len(dlbls) == 2

    def test_skips_missing_chart(self):
        """ChartPatch referencing a non-existent chart should not error."""
        xlsx_bytes = _make_xlsx_with_chart(num_points=3)
        patches = [
            ChartPatch(chart_index=999, pattern_fill_points=[0]),
        ]
        # Should not raise
        result = postprocess_xlsx(xlsx_bytes, patches)
        assert len(result) > 0

    def test_empty_patches_preserves_file(self):
        """An empty patches list should still produce a valid output."""
        xlsx_bytes = _make_xlsx_with_chart()
        result = postprocess_xlsx(xlsx_bytes, [])
        # Output should be approximately the same size
        assert abs(len(result) - len(xlsx_bytes)) < len(xlsx_bytes) * 0.5


# ---------------------------------------------------------------------------
# Tests: ChartPatch defaults
# ---------------------------------------------------------------------------

class TestChartPatch:
    def test_defaults(self):
        p = ChartPatch(chart_index=1)
        assert p.chart_index == 1
        assert p.pattern_fill_points == []
        assert p.asterisk_points == []
        assert p.series_index == 0

    def test_custom_values(self):
        p = ChartPatch(
            chart_index=3,
            pattern_fill_points=[1, 3],
            asterisk_points=[0, 2, 4],
            series_index=1,
        )
        assert p.chart_index == 3
        assert p.pattern_fill_points == [1, 3]
        assert p.asterisk_points == [0, 2, 4]
        assert p.series_index == 1
